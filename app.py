# app.py  — Flask web interface for PST importer + search
#
# Run:  .venv\Scripts\python.exe app.py
# Then open:  http://localhost:5000

import os
import sys
from dotenv import load_dotenv
load_dotenv()
import uuid
import json
import queue
import threading
import subprocess
import datetime

import logging
import logging.handlers
import mimetypes
import secrets
import bcrypt
from flask import Flask, render_template, request, jsonify, Response, stream_with_context, send_file, send_from_directory, redirect, url_for, session, flash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from authlib.integrations.flask_client import OAuth
from pymongo import MongoClient
import gridfs as gridfs_module
from bson import ObjectId
import io

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
MONGO_URI    = os.environ.get("MONGO_URI", "mongodb://127.0.0.1:27017/")
DB_NAME      = "pst_emails"
ADMIN_EMAILS = {"andy@computerhelpsos.com"}  # quota-exempt accounts
COLLECTION   = "pst_items"
UPLOAD_DIR   = os.path.join(os.path.dirname(__file__), "pst_files")
ATTACH_DIR    = os.path.join(os.path.dirname(__file__), "Attachments")
CSV_UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "csv_uploads")
PDF_DIR       = os.path.join(ATTACH_DIR, "pdf")
PDF_TEXT_DIR  = os.path.join(ATTACH_DIR, "pdf_text")    # text cache – PDFs
WORD_TEXT_DIR = os.path.join(ATTACH_DIR, "word_text")   # text cache – Word
EXCEL_TEXT_DIR= os.path.join(ATTACH_DIR, "excel_text")  # text cache – Excel
PPTX_TEXT_DIR = os.path.join(ATTACH_DIR, "pptx_text")   # text cache – PowerPoint
_INTERNAL_FOLDERS = {"pdf_text", "word_text", "excel_text", "pptx_text"}

# ---------------------------------------------------------------------------
# Auth config  — set these as environment variables in production
# ---------------------------------------------------------------------------
SECRET_KEY          = os.environ.get("SECRET_KEY",          secrets.token_hex(32))
GOOGLE_CLIENT_ID    = os.environ.get("GOOGLE_CLIENT_ID",    "")
GOOGLE_CLIENT_SECRET= os.environ.get("GOOGLE_CLIENT_SECRET","")
MS_CLIENT_ID        = os.environ.get("MS_CLIENT_ID",        "")
MS_CLIENT_SECRET    = os.environ.get("MS_CLIENT_SECRET",    "")
# Set to "1" to disable auth entirely (local dev)
AUTH_DISABLED       = os.environ.get("AUTH_DISABLED",       "0") == "1"

# ---------------------------------------------------------------------------
# Email config — for password-reset emails
# Resend is used when RESEND_API_KEY is set; falls back to SMTP otherwise.
# ---------------------------------------------------------------------------
RECAPTCHA_SITE_KEY   = os.environ.get("RECAPTCHA_SITE_KEY",   "6LcPOTctAAAAAEChuRQ4S4MRKjF9RxKxhZliKKbA")
RECAPTCHA_SECRET_KEY = os.environ.get("RECAPTCHA_SECRET_KEY", "6LcPOTctAAAAAGQ33pI3fOivpvOuo4LgG1TXtJQk")

RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
SMTP_HOST      = os.environ.get("SMTP_HOST",     "")
SMTP_PORT      = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER      = os.environ.get("SMTP_USER",     "")
SMTP_PASSWORD  = os.environ.get("SMTP_PASSWORD", "")
SMTP_FROM      = os.environ.get("SMTP_FROM",     SMTP_USER)
SMTP_USE_TLS   = os.environ.get("SMTP_USE_TLS",  "1") == "1"

# ---------------------------------------------------------------------------
# Stripe / Billing config  — set as environment variables in production
# ---------------------------------------------------------------------------
import stripe as _stripe_module

STRIPE_SECRET_KEY      = os.environ.get("STRIPE_SECRET_KEY",      "")
STRIPE_PUBLISHABLE_KEY = os.environ.get("STRIPE_PUBLISHABLE_KEY", "")
STRIPE_WEBHOOK_SECRET  = os.environ.get("STRIPE_WEBHOOK_SECRET",  "")

# One Stripe Price ID per paid plan — create these in your Stripe dashboard
STRIPE_PRICE_STARTER  = os.environ.get("STRIPE_PRICE_STARTER",  "")
STRIPE_PRICE_PRO      = os.environ.get("STRIPE_PRICE_PRO",      "")
STRIPE_PRICE_BUSINESS = os.environ.get("STRIPE_PRICE_BUSINESS", "")

# Plan definitions  {plan_key: {name, gb, monthly_usd, color, price_id}}
PLANS: dict = {
    "free":     {"name": "Free",     "gb": 1,   "monthly": 0,  "color": "secondary", "price_id": None},
    "starter":  {"name": "Starter",  "gb": 10,  "monthly": 9,  "color": "primary",   "price_id": STRIPE_PRICE_STARTER},
    "pro":      {"name": "Pro",      "gb": 50,  "monthly": 29, "color": "success",   "price_id": STRIPE_PRICE_PRO},
    "business": {"name": "Business", "gb": 200, "monthly": 79, "color": "warning",   "price_id": STRIPE_PRICE_BUSINESS},
}
os.makedirs(UPLOAD_DIR,    exist_ok=True)
os.makedirs(CSV_UPLOAD_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Session activity logger
# ---------------------------------------------------------------------------
_LOG_DIR = os.path.join(os.path.dirname(__file__), "logs")
os.makedirs(_LOG_DIR, exist_ok=True)

_activity_logger = logging.getLogger("session_activity")
_activity_logger.setLevel(logging.INFO)
_activity_logger.propagate = False
_log_handler = logging.handlers.TimedRotatingFileHandler(
    os.path.join(_LOG_DIR, "session_activity.log"),
    when="midnight", backupCount=90, encoding="utf-8",
)
_log_handler.setFormatter(logging.Formatter("%(asctime)s\t%(message)s",
                                             datefmt="%Y-%m-%d %H:%M:%S"))
_activity_logger.addHandler(_log_handler)

# Routes whose activity is too noisy to log individually
_LOG_SKIP_ENDPOINTS = frozenset({
    "static", "billing_usage", "progress", "build_index_route",
})

# Extensions that are never saved to disk or served, regardless of context
_BLOCKED_EXTENSIONS = frozenset({
    ".exe", ".bat", ".cmd", ".com", ".msi", ".dll", ".sys", ".drv",
    ".scr", ".pif", ".vbs", ".vbe", ".js",  ".jse", ".wsf", ".wsh",
    ".ps1", ".ps2", ".psm1", ".psd1", ".sh",  ".bash", ".zsh",
    ".jar", ".class", ".war", ".ear",
    ".hta", ".html5", ".xhtml",
    ".reg", ".inf", ".ins", ".isp",
    ".lnk", ".url", ".cpl", ".msc",
    ".iso", ".img", ".dmg",
    ".apk", ".ipa",
})

# Map file extensions → subfolder names under ATTACH_DIR
_EXT_FOLDER = {
    # PDF
    ".pdf":  "pdf",
    # Word
    ".doc":  "Word", ".docx": "Word", ".rtf": "Word", ".odt": "Word",
    # Excel
    ".xls":  "Excel", ".xlsx": "Excel", ".csv": "Excel",
    ".tsv":  "Excel", ".ods":  "Excel",
    # PowerPoint
    ".ppt":  "PowerPoint", ".pptx": "PowerPoint", ".odp": "PowerPoint",
    # Images
    ".jpg":  "Images", ".jpeg": "Images", ".png": "Images",
    ".gif":  "Images", ".bmp":  "Images", ".tiff": "Images",
    ".tif":  "Images", ".svg":  "Images", ".webp": "Images",
    # Videos
    ".mp4":  "Videos", ".avi":  "Videos", ".mov":  "Videos",
    ".wmv":  "Videos", ".mkv":  "Videos", ".flv":  "Videos",
    ".webm": "Videos", ".m4v":  "Videos", ".mpeg": "Videos",
    ".mpg":  "Videos", ".3gp":  "Videos",
    # Text / data
    ".txt":  "Text", ".log": "Text", ".xml": "Text", ".json": "Text",
    ".html": "Text", ".htm": "Text",
}

def _attach_folder(filename: str) -> str:
    """Return the per-user subfolder path (creating it if needed) for a given filename."""
    ext    = os.path.splitext(filename or "")[1].lower()
    if ext in _BLOCKED_EXTENSIONS:
        raise ValueError(f"File type '{ext}' is not allowed for security reasons.")
    subdir = _EXT_FOLDER.get(ext, "Other")
    path   = os.path.join(get_attach_dir(), subdir)
    os.makedirs(path, exist_ok=True)
    return path

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = None  # no size limit
app.config["SECRET_KEY"]         = SECRET_KEY

# ---------------------------------------------------------------------------
# Flask-Login
# ---------------------------------------------------------------------------
login_manager = LoginManager(app)
login_manager.login_view      = "login_page"
login_manager.login_message   = "Please sign in to continue."
login_manager.login_message_category = "info"

class User(UserMixin):
    def __init__(self, doc):
        self.id        = str(doc["_id"])
        self.email     = doc.get("email", "")
        self.name      = doc.get("name", "") or self.email
        self.provider  = doc.get("provider", "local")
        self.avatar_url= doc.get("avatar_url", "")

@login_manager.user_loader
def load_user(user_id):
    doc = get_users_col().find_one({"_id": user_id})
    return User(doc) if doc else None

# ---------------------------------------------------------------------------
# OAuth2 (Google + Microsoft)
# ---------------------------------------------------------------------------
oauth = OAuth(app)

oauth.register(
    name="google",
    client_id=GOOGLE_CLIENT_ID,
    client_secret=GOOGLE_CLIENT_SECRET,
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)

oauth.register(
    name="microsoft",
    client_id=MS_CLIENT_ID,
    client_secret=MS_CLIENT_SECRET,
    server_metadata_url="https://login.microsoftonline.com/common/v2.0/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)

# ---------------------------------------------------------------------------
# Auth gate  — runs before every request
# ---------------------------------------------------------------------------
_PUBLIC_ENDPOINTS = {
    "landing", "index",
    "login_page", "login_local", "register_page", "register_local",
    "auth_google", "auth_google_callback",
    "auth_microsoft", "auth_microsoft_callback",
    "billing_webhook",   # Stripe must reach this without a session cookie
    "forgot_password",
    "reset_password", "reset_password_submit",
    "static",
    "privacy_policy", "terms_of_service",
    "process_flow", "process_status",
}

@app.before_request
def require_login():
    if AUTH_DISABLED:
        return
    if current_user.is_authenticated:
        return
    if request.endpoint in _PUBLIC_ENDPOINTS:
        return
    return redirect(url_for("login_page", next=request.url))


@app.after_request
def log_activity(response):
    """Log each authenticated request to the daily session activity log."""
    endpoint = request.endpoint or ""
    if endpoint in _LOG_SKIP_ENDPOINTS:
        return response
    try:
        user_email = current_user.email if current_user.is_authenticated else "anonymous"
    except Exception:
        user_email = "anonymous"

    # Build a short description of what happened
    params = {}
    if request.args:
        params.update({k: v for k, v in request.args.items()
                       if k not in ("page",)})
    if request.is_json and request.method in ("POST", "PUT"):
        try:
            body = request.get_json(silent=True) or {}
            # include only safe, small keys
            for key in ("filename", "plan", "force"):
                if key in body:
                    params[key] = body[key]
        except Exception:
            pass

    detail = json.dumps(params, ensure_ascii=False) if params else ""
    _activity_logger.info(
        "%s\t%s\t%s %s\t%s\t%d",
        user_email,
        request.remote_addr or "-",
        request.method,
        request.path,
        detail,
        response.status_code,
    )
    return response

# In-memory job registry  {job_id: {"queue": Queue, "status": str, "filename": str}}
jobs: dict = {}


import re as _re_module

def get_user_id() -> str:
    """
    Return a filesystem- and MongoDB-safe identifier for the current user.
    Uses the logged-in user's email with non-alphanumeric chars replaced by '_'.
    Falls back to 'default' when auth is disabled or outside a request context.
    """
    try:
        if current_user and current_user.is_authenticated:
            safe = _re_module.sub(r"[^a-z0-9]", "_", current_user.email.lower()).strip("_")
            return safe or "default"
    except RuntimeError:
        pass  # outside request context (e.g. startup)
    return "default"


def _get_client():
    return MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)


def get_db():
    """Return the per-user MongoDB database."""
    uid = get_user_id()
    # Auth-disabled / anonymous → use the original DB_NAME unchanged
    db_name = DB_NAME if uid == "default" else f"{DB_NAME}_{uid}"
    return _get_client()[db_name]


def get_col():
    """Return the per-user pst_items collection."""
    return get_db()[COLLECTION]


def get_fs():
    """Return a per-user GridFS instance."""
    return gridfs_module.GridFS(get_db())


def get_users_col():
    """Users are always in the shared admin database."""
    client = _get_client()
    col    = client[DB_NAME]["pst_users"]
    col.create_index("email", unique=True, sparse=True)
    return col


# ── Per-user file-system directories ─────────────────────────────────────────

def get_upload_dir() -> str:
    """Per-user PST upload directory."""
    path = os.path.join(UPLOAD_DIR, get_user_id())
    os.makedirs(path, exist_ok=True)
    return path


def get_attach_dir() -> str:
    """Per-user attachments root directory."""
    path = os.path.join(ATTACH_DIR, get_user_id())
    os.makedirs(path, exist_ok=True)
    return path


def _user_attach_subdir(subdir: str) -> str:
    path = os.path.join(get_attach_dir(), subdir)
    os.makedirs(path, exist_ok=True)
    return path


def get_pdf_dir()        -> str: return _user_attach_subdir("pdf")
def get_pdf_text_dir()   -> str: return _user_attach_subdir("pdf_text")
def get_word_text_dir()  -> str: return _user_attach_subdir("word_text")
def get_excel_text_dir() -> str: return _user_attach_subdir("excel_text")
def get_pptx_text_dir()  -> str: return _user_attach_subdir("pptx_text")


# ── Billing helpers ───────────────────────────────────────────────────────────

def get_user_plan_name() -> str:
    """Return the current user's plan key (e.g. 'free', 'pro')."""
    try:
        if not current_user.is_authenticated:
            return "free"
        doc = get_users_col().find_one({"_id": current_user.id}) or {}
        return doc.get("plan", "free")
    except Exception:
        return "free"


def get_user_quota_bytes() -> int:
    plan = PLANS.get(get_user_plan_name(), PLANS["free"])
    return plan["gb"] * 1024 ** 3


def is_admin() -> bool:
    """Return True if the current user is quota-exempt."""
    try:
        if current_user and current_user.is_authenticated:
            return current_user.email.lower() in {e.lower() for e in ADMIN_EMAILS}
    except Exception:
        pass
    return False


def get_user_storage_bytes() -> int:
    """Sum disk usage across the user's PST files and attachments.

    Returns 0 when the user has no records in the database (orphaned
    disk files don't count toward the displayed quota).
    """
    try:
        if get_col().estimated_document_count() == 0:
            return 0
    except Exception:
        pass
    total = 0
    for base in [get_upload_dir(), get_attach_dir()]:
        if os.path.isdir(base):
            for dirpath, _, filenames in os.walk(base):
                for fname in filenames:
                    try:
                        total += os.path.getsize(os.path.join(dirpath, fname))
                    except Exception:
                        pass
    return total


def _get_or_create_stripe_customer(user_doc: dict) -> str:
    """Return existing Stripe customer ID or create a new one."""
    if not STRIPE_SECRET_KEY:
        return ""
    _stripe_module.api_key = STRIPE_SECRET_KEY
    cid = user_doc.get("stripe_customer_id")
    if cid:
        return cid
    customer = _stripe_module.Customer.create(
        email=user_doc.get("email", ""),
        name=user_doc.get("name", ""),
        metadata={"user_id": str(user_doc.get("_id", ""))},
    )
    get_users_col().update_one(
        {"_id": user_doc["_id"]},
        {"$set": {"stripe_customer_id": customer.id}},
    )
    return customer.id


def _apply_stripe_subscription(customer_id: str, subscription):
    """Update user plan in DB based on a Stripe subscription object."""
    users = get_users_col()
    try:
        price_id = subscription["items"]["data"][0]["price"]["id"]
    except (KeyError, IndexError):
        price_id = ""
    plan_name = next(
        (k for k, v in PLANS.items() if v.get("price_id") and v["price_id"] == price_id),
        "free",
    )
    users.update_one(
        {"stripe_customer_id": customer_id},
        {"$set": {
            "plan":                    plan_name,
            "plan_status":             subscription.get("status", "active"),
            "stripe_subscription_id":  subscription.get("id"),
        }},
    )


# ---------------------------------------------------------------------------
# Billing Routes
# ---------------------------------------------------------------------------

@app.route("/billing")
def billing_page():
    user_doc     = get_users_col().find_one({"_id": current_user.id}) or {}
    plan_name    = user_doc.get("plan", "free")
    plan         = PLANS.get(plan_name, PLANS["free"])
    plan_status  = user_doc.get("plan_status", "active")
    storage_bytes= get_user_storage_bytes()
    quota_bytes  = plan["gb"] * 1024 ** 3
    storage_pct  = min(100, int(storage_bytes / max(quota_bytes, 1) * 100))

    def _fmt(b):
        if b < 1_048_576:     return f"{b/1_024:.1f} KB"
        if b < 1_073_741_824: return f"{b/1_048_576:.1f} MB"
        return f"{b/1_073_741_824:.2f} GB"

    return render_template(
        "billing.html",
        plan_name       = plan_name,
        plan            = plan,
        plan_status     = plan_status,
        plans           = PLANS,
        storage_bytes   = storage_bytes,
        storage_str     = _fmt(storage_bytes),
        quota_bytes     = quota_bytes,
        quota_str       = f"{plan['gb']} GB",
        storage_pct     = storage_pct,
        stripe_pk       = STRIPE_PUBLISHABLE_KEY,
        stripe_configured = bool(STRIPE_SECRET_KEY),
        has_subscription  = bool(user_doc.get("stripe_subscription_id")),
    )


@app.route("/billing/create-checkout-session", methods=["POST"])
def billing_create_checkout():
    if not STRIPE_SECRET_KEY:
        flash("Stripe is not configured on this server yet.", "warning")
        return redirect(url_for("billing_page"))
    _stripe_module.api_key = STRIPE_SECRET_KEY

    plan_name = request.form.get("plan", "")
    plan      = PLANS.get(plan_name)
    if not plan or not plan.get("price_id"):
        flash("Invalid plan selected.", "danger")
        return redirect(url_for("billing_page"))

    user_doc    = get_users_col().find_one({"_id": current_user.id}) or {
        "_id": current_user.id, "email": current_user.email, "name": current_user.name
    }
    customer_id = _get_or_create_stripe_customer(user_doc)

    session = _stripe_module.checkout.Session.create(
        customer              = customer_id,
        payment_method_types  = ["card"],
        line_items            = [{"price": plan["price_id"], "quantity": 1}],
        mode                  = "subscription",
        allow_promotion_codes = True,
        success_url = url_for("billing_success", _external=True) + "?session_id={CHECKOUT_SESSION_ID}",
        cancel_url  = url_for("billing_page", _external=True),
    )
    return redirect(session.url, code=303)


@app.route("/billing/success")
def billing_success():
    flash("🎉 Subscription activated! Your storage quota has been upgraded.", "success")
    return redirect(url_for("billing_page"))


STRIPE_PRICE_VIP = os.environ.get("STRIPE_PRICE_VIP", "price_1ToWBkDMYOvxQd9SY9NbBdum")

@app.route("/billing/vip-checkout", methods=["POST"])
def billing_vip_checkout():
    if not STRIPE_SECRET_KEY:
        flash("Stripe is not configured on this server yet.", "warning")
        return redirect(url_for("billing_page"))
    _stripe_module.api_key = STRIPE_SECRET_KEY

    first_name = request.form.get("first_name", "").strip()
    last_name  = request.form.get("last_name",  "").strip()
    phone      = request.form.get("phone",      "").strip()
    email      = request.form.get("email",      "").strip()
    best_time  = request.form.get("best_time",  "").strip()

    # Store the enquiry in MongoDB so it survives whether they complete payment
    _get_client()["pst_emails_admin"]["vip_orders"].insert_one({
        "first_name": first_name, "last_name": last_name,
        "phone": phone, "email": email, "best_time": best_time,
        "status": "pending_payment",
        "submitted_at": datetime.datetime.utcnow(),
    })

    # Send admin notification email
    subject    = f"[PST Browser] New VIP Setup request — {first_name} {last_name}"
    body_plain = (f"Name:       {first_name} {last_name}\n"
                  f"Email:      {email}\n"
                  f"Phone:      {phone}\n"
                  f"Best time:  {best_time}\n\n"
                  f"Payment pending — they are being redirected to Stripe now.")
    body_html  = (f"<p><strong>Name:</strong> {first_name} {last_name}<br>"
                  f"<strong>Email:</strong> {email}<br>"
                  f"<strong>Phone:</strong> {phone}<br>"
                  f"<strong>Best time to call:</strong> {best_time}</p>"
                  f"<p><em>Payment pending — redirecting to Stripe now.</em></p>")
    _send_notification_email("andy@computerhelpsos.com", subject, body_plain, body_html)

    # Create Stripe checkout session (one-time payment)
    user_doc    = get_users_col().find_one({"_id": current_user.id}) or {
        "_id": current_user.id, "email": current_user.email, "name": current_user.name
    }
    customer_id = _get_or_create_stripe_customer(user_doc)

    session = _stripe_module.checkout.Session.create(
        customer             = customer_id,
        payment_method_types = ["card"],
        line_items           = [{"price": STRIPE_PRICE_VIP, "quantity": 1}],
        mode                 = "payment",
        allow_promotion_codes= True,
        metadata             = {"type": "vip_setup", "email": email, "phone": phone, "best_time": best_time},
        success_url = url_for("billing_vip_success", _external=True),
        cancel_url  = url_for("billing_page", _external=True),
    )
    return redirect(session.url, code=303)


@app.route("/billing/vip-success")
def billing_vip_success():
    # Send confirmation email to admin with "payment complete"
    subject    = "[PST Browser] VIP Setup — payment confirmed"
    body_plain = "A VIP Setup payment has been completed. Check the vip_orders collection for details."
    body_html  = "<p>A VIP Setup payment has been completed. Check <strong>vip_orders</strong> in MongoDB for details.</p>"
    _send_notification_email("andy@computerhelpsos.com", subject, body_plain, body_html)
    flash("🎉 Payment received! We'll contact you shortly to schedule your remote session.", "success")
    return redirect(url_for("billing_page"))


@app.route("/billing/portal", methods=["POST"])
def billing_portal():
    if not STRIPE_SECRET_KEY:
        flash("Stripe is not configured on this server yet.", "warning")
        return redirect(url_for("billing_page"))
    _stripe_module.api_key = STRIPE_SECRET_KEY

    user_doc    = get_users_col().find_one({"_id": current_user.id}) or {}
    customer_id = user_doc.get("stripe_customer_id")
    if not customer_id:
        flash("No billing account found — please subscribe first.", "warning")
        return redirect(url_for("billing_page"))

    portal = _stripe_module.billing_portal.Session.create(
        customer   = customer_id,
        return_url = url_for("billing_page", _external=True),
    )
    return redirect(portal.url, code=303)


@app.route("/billing/usage")
def billing_usage():
    """Quick JSON endpoint for the navbar storage bar."""
    plan_name     = get_user_plan_name()
    plan          = PLANS.get(plan_name, PLANS["free"])
    storage_bytes = get_user_storage_bytes()
    quota_bytes   = plan["gb"] * 1024 ** 3
    return jsonify({
        "plan":          plan_name,
        "plan_name":     plan["name"],
        "storage_bytes": storage_bytes,
        "quota_bytes":   quota_bytes,
        "storage_pct":   min(100, int(storage_bytes / max(quota_bytes, 1) * 100)),
        "storage_gb":    round(storage_bytes / 1024 ** 3, 2),
        "quota_gb":      plan["gb"],
    })


@app.route("/billing/webhook", methods=["POST"])
def billing_webhook():
    """Stripe sends events here — no auth cookie required."""
    if not STRIPE_SECRET_KEY:
        return jsonify({"error": "Stripe not configured"}), 400
    _stripe_module.api_key = STRIPE_SECRET_KEY

    payload = request.get_data()
    sig     = request.headers.get("Stripe-Signature", "")
    try:
        event = _stripe_module.Webhook.construct_event(
            payload, sig, STRIPE_WEBHOOK_SECRET
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    obj  = event["data"]["object"]
    etype= event["type"]

    def _sget(o, key, default=None):
        try:
            return o[key]
        except (KeyError, TypeError):
            return default

    if etype == "checkout.session.completed":
        sub_id = _sget(obj, "subscription")
        cid    = _sget(obj, "customer")
        if sub_id and cid:
            sub = _stripe_module.Subscription.retrieve(sub_id)
            _apply_stripe_subscription(cid, sub)

    elif etype in ("customer.subscription.updated",
                   "customer.subscription.created"):
        _apply_stripe_subscription(_sget(obj, "customer", ""), obj)

    elif etype == "customer.subscription.deleted":
        cid = _sget(obj, "customer", "")
        if cid:
            get_users_col().update_one(
                {"stripe_customer_id": cid},
                {"$set": {"plan": "free", "plan_status": "canceled",
                           "stripe_subscription_id": None}},
            )

    elif etype == "invoice.payment_failed":
        cid = _sget(obj, "customer", "")
        if cid:
            get_users_col().update_one(
                {"stripe_customer_id": cid},
                {"$set": {"plan_status": "past_due"}},
            )

    return jsonify({"received": True})


# ---------------------------------------------------------------------------
# Auth Routes
# ---------------------------------------------------------------------------

def _send_welcome_email(to_addr: str, name: str):
    subject    = "Welcome to PSTBrowser!"
    body_plain = (
        f"Hi {name},\n\n"
        f"Welcome to PSTBrowser! We're glad to have you.\n\n"
        f"You can start importing your PST file right away at https://pstbrowser.com/?tab=import\n\n"
        f"For support please email support@pstbrowser.com or call (571) 200-1551.\n\n"
        f"Thank you,\nPSTBrowser"
    )
    body_html = (
        f"<p>Hi {name},</p>"
        f"<p>Welcome to PSTBrowser! We're glad to have you.</p>"
        f"<p>You can start importing your PST file right away at "
        f"<a href='https://pstbrowser.com/?tab=import'>pstbrowser.com</a>.</p>"
        f"<p>For support please email "
        f"<a href='mailto:support@pstbrowser.com'>support@pstbrowser.com</a> "
        f"or call (571) 200-1551.</p>"
        f"<p>Thank you,<br>PSTBrowser</p>"
    )
    _send_notification_email(to_addr, subject, body_plain, body_html)


def _send_notification_email(to_addr: str, subject: str, body_plain: str, body_html: str):
    """Send a notification email using Resend or SMTP (same infrastructure as password reset)."""
    if not to_addr:
        return
    try:
        if RESEND_API_KEY:
            import resend as _resend
            _resend.api_key = RESEND_API_KEY
            from_addr = SMTP_FROM or "onboarding@resend.dev"
            _resend.Emails.send({
                "from": from_addr, "to": [to_addr],
                "subject": subject, "text": body_plain, "html": body_html,
            })
            return
        if SMTP_HOST and SMTP_USER:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text      import MIMEText
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"]    = SMTP_FROM or SMTP_USER
            msg["To"]      = to_addr
            msg.attach(MIMEText(body_plain, "plain", "utf-8"))
            msg.attach(MIMEText(body_html,  "html",  "utf-8"))
            if SMTP_USE_TLS:
                server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10)
                server.ehlo(); server.starttls()
            else:
                server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=10)
            if SMTP_USER and SMTP_PASSWORD:
                server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM or SMTP_USER, [to_addr], msg.as_bytes())
            server.quit()
    except Exception as e:
        print(f"[EMAIL ERROR] Notification email to {to_addr} failed: {e}", flush=True)
        app.logger.error("Notification email failed: %s", e)


def _verify_recaptcha(token: str, min_score: float = 0.5) -> bool:
    """Return True if the reCAPTCHA v3 token is valid and score >= min_score."""
    if not RECAPTCHA_SECRET_KEY or not token:
        return True
    import urllib.request
    import urllib.parse
    data = urllib.parse.urlencode({
        "secret":   RECAPTCHA_SECRET_KEY,
        "response": token,
        "remoteip": request.remote_addr or "",
    }).encode()
    try:
        with urllib.request.urlopen(
            "https://www.google.com/recaptcha/api/siteverify", data, timeout=5
        ) as resp:
            result = json.loads(resp.read())
        return bool(result.get("success")) and float(result.get("score", 0)) >= min_score
    except Exception:
        return False


@app.route("/privacy")
def privacy_policy():
    return render_template("privacy.html")


@app.route("/terms")
def terms_of_service():
    return render_template("terms.html")


@app.route("/suggestions", methods=["GET", "POST"])
def suggestions():
    if request.method == "POST":
        name     = request.form.get("name", "").strip()
        email    = request.form.get("email", "").strip()
        category = request.form.get("category", "Other").strip()
        body     = request.form.get("body", "").strip()
        if not name or not email or not body:
            flash("Please fill in all fields.", "warning")
            return redirect(url_for("suggestions"))
        doc = {
            "name": name, "email": email,
            "category": category, "body": body,
            "submitted_at": datetime.datetime.utcnow(),
        }
        try:
            _get_client()["pst_emails_admin"]["suggestions"].insert_one(doc)
        except Exception as e:
            print(f"[SUGGESTIONS] DB error: {e}", flush=True)
        # Email notification to admin
        subject    = f"[PST Browser] New suggestion from {name}"
        body_plain = f"From: {name} <{email}>\nCategory: {category}\n\n{body}"
        body_html  = (f"<p><strong>From:</strong> {name} &lt;{email}&gt;<br>"
                      f"<strong>Category:</strong> {category}</p>"
                      f"<p style='white-space:pre-wrap'>{body}</p>")
        _send_notification_email("andy@computerhelpsos.com", subject, body_plain, body_html)
        flash("Thank you! Your suggestion has been submitted.", "success")
        return redirect(url_for("suggestions"))
    return render_template("suggestions.html")


@app.route("/admin/accounts")
def admin_accounts():
    if not (current_user.is_authenticated and current_user.email.lower() in {e.lower() for e in ADMIN_EMAILS}):
        return redirect(url_for("login_page"))
    users = list(get_users_col().find({}, {"_id":0,"email":1,"name":1,"provider":1,"created_at":1,"last_login":1})
                 .sort("created_at", -1))
    return render_template("admin_accounts.html", users=users)


@app.route("/admin/suggestions")
def admin_suggestions():
    if not (current_user.is_authenticated and current_user.email.lower() in {e.lower() for e in ADMIN_EMAILS}):
        return redirect(url_for("login_page"))
    docs = list(_get_client()["pst_emails_admin"]["suggestions"]
                .find({}).sort("submitted_at", -1))
    # convert ObjectId to string for template use
    for d in docs:
        d["_id"] = str(d["_id"])
    return render_template("admin_suggestions.html", suggestions=docs)


@app.route("/admin/suggestions/<sid>/status", methods=["POST"])
def admin_suggestion_status(sid):
    if not (current_user.is_authenticated and current_user.email.lower() in {e.lower() for e in ADMIN_EMAILS}):
        return jsonify({"error": "forbidden"}), 403
    from bson import ObjectId
    status = request.json.get("status", "new")
    _get_client()["pst_emails_admin"]["suggestions"].update_one(
        {"_id": ObjectId(sid)}, {"$set": {"status": status}})
    return jsonify({"ok": True})


@app.route("/admin/suggestions/<sid>/delete", methods=["POST"])
def admin_suggestion_delete(sid):
    if not (current_user.is_authenticated and current_user.email.lower() in {e.lower() for e in ADMIN_EMAILS}):
        return jsonify({"error": "forbidden"}), 403
    from bson import ObjectId
    _get_client()["pst_emails_admin"]["suggestions"].delete_one({"_id": ObjectId(sid)})
    return jsonify({"ok": True})


@app.route("/auth/login")
def login_page():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    return render_template("login.html",
                           google_enabled=bool(GOOGLE_CLIENT_ID),
                           ms_enabled=bool(MS_CLIENT_ID),
                           recaptcha_site_key=RECAPTCHA_SITE_KEY)


@app.route("/auth/login", methods=["POST"])
def login_local():
    if not _verify_recaptcha(request.form.get("g-recaptcha-response", "")):
        flash("Please complete the CAPTCHA.", "danger")
        return redirect(url_for("login_page"))
    email    = (request.form.get("email")    or "").strip().lower()
    password = (request.form.get("password") or "").strip()
    if not email or not password:
        flash("Email and password are required.", "danger")
        return redirect(url_for("login_page"))
    doc = get_users_col().find_one({"_id": email, "provider": "local"})
    if not doc or not bcrypt.checkpw(password.encode(), doc["password_hash"]):
        flash("Invalid email or password.", "danger")
        return redirect(url_for("login_page"))
    get_users_col().update_one({"_id": email}, {"$set": {"last_login": datetime.datetime.utcnow()}})
    login_user(User(doc), remember=True)
    next_url = request.args.get("next") or url_for("index")
    return redirect(next_url)


@app.route("/auth/register")
def register_page():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    return render_template("login.html", show_register=True,
                           google_enabled=bool(GOOGLE_CLIENT_ID),
                           ms_enabled=bool(MS_CLIENT_ID),
                           recaptcha_site_key=RECAPTCHA_SITE_KEY)


@app.route("/auth/register", methods=["POST"])
def register_local():
    if not _verify_recaptcha(request.form.get("g-recaptcha-response", "")):
        flash("Please complete the CAPTCHA.", "danger")
        return redirect(url_for("register_page"))
    name     = (request.form.get("name")     or "").strip()
    email    = (request.form.get("email")    or "").strip().lower()
    password = (request.form.get("password") or "").strip()
    if not email or not password or not name:
        flash("All fields are required.", "danger")
        return redirect(url_for("register_page"))
    if len(password) < 8:
        flash("Password must be at least 8 characters.", "danger")
        return redirect(url_for("register_page"))
    users = get_users_col()
    if users.find_one({"_id": email}):
        flash("An account with that email already exists.", "warning")
        return redirect(url_for("login_page"))
    pw_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt())
    doc = {
        "_id":           email,
        "email":         email,
        "name":          name,
        "provider":      "local",
        "password_hash": pw_hash,
        "avatar_url":    "",
        "created_at":    datetime.datetime.utcnow(),
        "last_login":    datetime.datetime.utcnow(),
    }
    users.insert_one(doc)
    login_user(User(doc), remember=True)
    _send_welcome_email(email, name)
    flash(f"Welcome, {name}! Your account has been created.", "success")
    return redirect(url_for("index") + "?tab=import")


@app.route("/auth/change-password", methods=["POST"])
def change_password():
    if not current_user.is_authenticated:
        return jsonify({"error": "Not authenticated"}), 401
    if current_user.provider != "local":
        return jsonify({"error": "Password change is only available for local accounts."}), 400

    data         = request.get_json(force=True)
    current_pw   = (data.get("current_password") or "").strip()
    new_pw       = (data.get("new_password")     or "").strip()
    confirm_pw   = (data.get("confirm_password") or "").strip()

    if not current_pw or not new_pw or not confirm_pw:
        return jsonify({"error": "All fields are required."}), 400
    if new_pw != confirm_pw:
        return jsonify({"error": "New passwords do not match."}), 400
    if len(new_pw) < 8:
        return jsonify({"error": "New password must be at least 8 characters."}), 400

    users = get_users_col()
    doc   = users.find_one({"_id": current_user.id, "provider": "local"})
    if not doc or not bcrypt.checkpw(current_pw.encode(), doc["password_hash"]):
        return jsonify({"error": "Current password is incorrect."}), 403

    new_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt())
    users.update_one({"_id": current_user.id}, {"$set": {"password_hash": new_hash}})
    return jsonify({"ok": True})


def _send_reset_email(to_addr: str, reset_url: str):
    """Send a password-reset link. Uses Resend when RESEND_API_KEY is set, else SMTP."""
    subject = "Reset your PST Browser password"
    body_plain = (
        f"Hi,\n\n"
        f"We received a request to reset the password for your account ({to_addr}).\n\n"
        f"Click the link below to choose a new password (valid for 1 hour):\n\n"
        f"{reset_url}\n\n"
        f"If you didn't request this, you can safely ignore this email.\n\n"
        f"— PST Browser"
    )
    body_html = (
        f"<p>Hi,</p>"
        f"<p>We received a request to reset the password for your account "
        f"(<strong>{to_addr}</strong>).</p>"
        f"<p><a href=\"{reset_url}\" style=\"background:#2563eb;color:#fff;"
        f"padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:600;"
        f"display:inline-block\">Reset password</a></p>"
        f"<p style=\"color:#6b7280;font-size:.9em\">Link expires in 1 hour. "
        f"If you didn't request this, ignore this email.</p>"
        f"<p style=\"color:#6b7280;font-size:.85em\">Or copy this URL: {reset_url}</p>"
    )

    if RESEND_API_KEY:
        import resend as _resend
        _resend.api_key = RESEND_API_KEY
        from_addr = SMTP_FROM or "onboarding@resend.dev"
        _resend.Emails.send({
            "from":    from_addr,
            "to":      [to_addr],
            "subject": subject,
            "text":    body_plain,
            "html":    body_html,
        })
        return

    # SMTP fallback
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = SMTP_FROM or SMTP_USER
    msg["To"]      = to_addr
    msg.attach(MIMEText(body_plain, "plain", "utf-8"))
    msg.attach(MIMEText(body_html,  "html",  "utf-8"))
    if SMTP_USE_TLS:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10)
        server.ehlo()
        server.starttls()
    else:
        server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=10)
    if SMTP_USER and SMTP_PASSWORD:
        server.login(SMTP_USER, SMTP_PASSWORD)
    server.sendmail(SMTP_FROM or SMTP_USER, [to_addr], msg.as_bytes())
    server.quit()


@app.route("/auth/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "GET":
        return render_template("forgot_password.html")

    email = (request.form.get("email") or "").strip().lower()
    if not email:
        flash("Please enter your email address.", "danger")
        return redirect(url_for("forgot_password"))

    # Always show the same message to avoid user enumeration
    flash("If that address is registered, you'll receive a reset link shortly.", "info")

    users = get_users_col()
    doc   = users.find_one({"_id": email, "provider": "local"})
    if not doc:
        return redirect(url_for("login_page"))

    token      = secrets.token_urlsafe(32)
    expires_at = datetime.datetime.utcnow() + datetime.timedelta(hours=1)
    users.update_one(
        {"_id": email},
        {"$set": {"pw_reset_token": token, "pw_reset_expires": expires_at}},
    )

    reset_url = url_for("reset_password", token=token, _external=True)
    try:
        _send_reset_email(email, reset_url)
    except Exception as e:
        app.logger.error("Password-reset email failed: %s", e)
        # Fail silently to the user — avoid leaking SMTP errors

    return redirect(url_for("login_page"))


@app.route("/auth/reset-password/<token>", methods=["GET"])
def reset_password(token: str):
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    users = get_users_col()
    doc   = users.find_one({"pw_reset_token": token})
    if not doc or doc.get("pw_reset_expires", datetime.datetime.min) < datetime.datetime.utcnow():
        flash("This password-reset link is invalid or has expired.", "danger")
        return redirect(url_for("forgot_password"))
    return render_template("reset_password.html", token=token)


@app.route("/auth/reset-password/<token>", methods=["POST"])
def reset_password_submit(token: str):
    new_pw     = (request.form.get("password")         or "").strip()
    confirm_pw = (request.form.get("confirm_password") or "").strip()

    if not new_pw or len(new_pw) < 8:
        flash("Password must be at least 8 characters.", "danger")
        return redirect(url_for("reset_password", token=token))
    if new_pw != confirm_pw:
        flash("Passwords do not match.", "danger")
        return redirect(url_for("reset_password", token=token))

    users = get_users_col()
    doc   = users.find_one({"pw_reset_token": token})
    if not doc or doc.get("pw_reset_expires", datetime.datetime.min) < datetime.datetime.utcnow():
        flash("This password-reset link is invalid or has expired.", "danger")
        return redirect(url_for("forgot_password"))

    new_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt())
    users.update_one(
        {"_id": doc["_id"]},
        {"$set":   {"password_hash": new_hash},
         "$unset": {"pw_reset_token": "", "pw_reset_expires": ""}},
    )
    flash("Your password has been reset. Please sign in.", "success")
    return redirect(url_for("login_page"))


@app.route("/auth/logout")
def logout():
    logout_user()
    flash("You have been signed out.", "info")
    return redirect(url_for("login_page"))


# ── Google OAuth ──────────────────────────────────────────────────────────────

@app.route("/auth/google")
def auth_google():
    if not GOOGLE_CLIENT_ID:
        flash("Google sign-in is not configured.", "warning")
        return redirect(url_for("login_page"))
    redirect_uri = url_for("auth_google_callback", _external=True)
    return oauth.google.authorize_redirect(redirect_uri)


@app.route("/auth/google/callback")
def auth_google_callback():
    try:
        token    = oauth.google.authorize_access_token()
        userinfo = token.get("userinfo") or oauth.google.userinfo()
    except Exception as e:
        flash(f"Google sign-in failed: {e}", "danger")
        return redirect(url_for("login_page"))

    email      = (userinfo.get("email") or "").lower()
    name       = userinfo.get("name")  or email
    avatar_url = userinfo.get("picture") or ""
    if not email:
        flash("Could not retrieve your email from Google.", "danger")
        return redirect(url_for("login_page"))

    users = get_users_col()
    doc   = users.find_one({"_id": email})
    now   = datetime.datetime.utcnow()
    is_new = doc is None
    if doc:
        users.update_one({"_id": email}, {"$set": {"name": name, "avatar_url": avatar_url, "last_login": now}})
        doc.update({"name": name, "avatar_url": avatar_url})
    else:
        doc = {"_id": email, "email": email, "name": name, "provider": "google",
               "avatar_url": avatar_url, "created_at": now, "last_login": now}
        users.insert_one(doc)
        _send_welcome_email(email, name)

    login_user(User(doc), remember=True)
    default = url_for("index") + ("?tab=import" if is_new else "")
    return redirect(request.args.get("next") or default)


# ── Microsoft OAuth ───────────────────────────────────────────────────────────

@app.route("/auth/microsoft")
def auth_microsoft():
    if not MS_CLIENT_ID:
        flash("Microsoft sign-in is not configured.", "warning")
        return redirect(url_for("login_page"))
    redirect_uri = url_for("auth_microsoft_callback", _external=True)
    return oauth.microsoft.authorize_redirect(redirect_uri)


@app.route("/auth/microsoft/callback")
def auth_microsoft_callback():
    try:
        token    = oauth.microsoft.authorize_access_token(
            claims_options={"iss": {"essential": False}}
        )
        userinfo = token.get("userinfo") or {}
        if not userinfo:
            import requests as _req
            resp     = _req.get("https://graph.microsoft.com/v1.0/me",
                                headers={"Authorization": f"Bearer {token['access_token']}"})
            ms_data  = resp.json()
            userinfo = {"email": ms_data.get("mail") or ms_data.get("userPrincipalName", ""),
                        "name":  ms_data.get("displayName", "")}
    except Exception as e:
        flash(f"Microsoft sign-in failed: {e}", "danger")
        return redirect(url_for("login_page"))

    email      = (userinfo.get("email") or "").lower()
    name       = userinfo.get("name")  or email
    avatar_url = ""
    if not email:
        flash("Could not retrieve your email from Microsoft.", "danger")
        return redirect(url_for("login_page"))

    users = get_users_col()
    doc   = users.find_one({"_id": email})
    now   = datetime.datetime.utcnow()
    is_new = doc is None
    if doc:
        users.update_one({"_id": email}, {"$set": {"name": name, "last_login": now}})
        doc.update({"name": name})
    else:
        doc = {"_id": email, "email": email, "name": name, "provider": "microsoft",
               "avatar_url": avatar_url, "created_at": now, "last_login": now}
        users.insert_one(doc)
        _send_welcome_email(email, name)

    login_user(User(doc), remember=True)
    default = url_for("index") + ("?tab=import" if is_new else "")
    return redirect(request.args.get("next") or default)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/landing")
def landing():
    return render_template("landing.html")

@app.route("/")
def index():
    if not current_user.is_authenticated:
        return render_template("landing.html")
    return render_template("index.html")


# ---------------------------------------------------------------------------
# Browse Attachments
# ---------------------------------------------------------------------------

_ATTACH_ICONS = {
    "pdf": "bi-file-earmark-pdf text-danger",
    "Word": "bi-file-earmark-word text-primary",
    "Excel": "bi-file-earmark-excel text-success",
    "PowerPoint": "bi-file-earmark-ppt text-warning",
    "Images": "bi-image text-info",
    "Videos": "bi-camera-video text-purple",
    "Text": "bi-file-earmark-text text-secondary",
    "Other": "bi-file-earmark text-muted",
}

@app.route("/attachments")
def browse_attachments():
    attach_dir = get_attach_dir()
    folder     = request.args.get("folder", "")
    q          = request.args.get("q", "").strip().lower()
    page       = max(1, int(request.args.get("page", 1)))
    sort       = request.args.get("sort", "name")
    order      = request.args.get("order", "asc")
    per_page   = 50

    # Build folder list with counts
    folders = []
    for name in ["pdf", "Word", "Excel", "PowerPoint", "Images", "Videos", "Text", "Other"]:
        path = os.path.join(attach_dir, name)
        if os.path.isdir(path):
            count = sum(1 for f in os.listdir(path)
                        if os.path.isfile(os.path.join(path, f)))
            if count:
                folders.append({"name": name, "count": count,
                                 "icon": _ATTACH_ICONS.get(name, "bi-folder")})

    # Default to first folder
    if not folder and folders:
        folder = folders[0]["name"]

    # List files — search all folders when query is present, else selected folder only
    files = []
    search_folders = [f["name"] for f in folders] if q else ([folder] if folder else [])
    for search_folder in search_folders:
        folder_path = os.path.join(attach_dir, search_folder)
        if not os.path.isdir(folder_path):
            continue
        for fname in os.listdir(folder_path):
            fpath = os.path.join(folder_path, fname)
            if not os.path.isfile(fpath):
                continue
            if q and q not in fname.lower():
                continue
            stat = os.stat(fpath)
            files.append({
                "name":     fname,
                "folder":   search_folder,
                "size":     stat.st_size,
                "modified": datetime.datetime.fromtimestamp(stat.st_mtime),
            })

    # Sort
    sort_key = {"name": "name", "size": "size", "modified": "modified", "type": "folder"}.get(sort, "name")
    files.sort(key=lambda f: f[sort_key], reverse=(order == "desc"))

    total = len(files)
    files = files[(page - 1) * per_page: page * per_page]
    pages = max(1, -(-total // per_page))

    def fmt_size(b):
        if b < 1024:        return f"{b} B"
        if b < 1048576:     return f"{b/1024:.1f} KB"
        if b < 1073741824:  return f"{b/1048576:.1f} MB"
        return f"{b/1073741824:.2f} GB"

    return render_template("browse_attachments.html",
                           folders=folders, folder=folder,
                           files=files, total=total,
                           page=page, pages=pages, per_page=per_page,
                           q=q, fmt_size=fmt_size,
                           icons=_ATTACH_ICONS,
                           sort=sort, order=order)


@app.route("/attachments/search-names")
def search_attachment_names():
    """Return files whose names contain the query string."""
    q = request.args.get("q", "").strip().lower()
    if not q:
        return jsonify({"results": []})
    attach_dir = get_attach_dir()
    results = []
    for folder_name in ["pdf", "Word", "Excel", "PowerPoint", "Images", "Videos", "Text", "Other"]:
        folder_path = os.path.join(attach_dir, folder_name)
        if not os.path.isdir(folder_path):
            continue
        for fname in os.listdir(folder_path):
            if q in fname.lower() and os.path.isfile(os.path.join(folder_path, fname)):
                results.append({"folder": folder_name, "filename": fname})
                if len(results) >= 200:
                    break
    return jsonify({"results": results})


@app.route("/attachments/download")
def download_attachment_file():
    folder   = request.args.get("folder", "")
    filename = request.args.get("file", "")
    # Security: no path traversal
    if not folder or not filename or ".." in folder or ".." in filename:
        return "Invalid path", 400
    ext = os.path.splitext(filename)[1].lower()
    if ext in _BLOCKED_EXTENSIONS:
        return "File type not allowed", 403
    path = os.path.join(get_attach_dir(), os.path.basename(folder),
                        os.path.basename(filename))
    if not os.path.isfile(path):
        return "Not found", 404
    return send_file(path, as_attachment=True, download_name=filename)


# ---------------------------------------------------------------------------
# Clear My Data
# ---------------------------------------------------------------------------

@app.route("/account/clear-data", methods=["POST"])
def clear_user_data():
    """Delete all files and database records for the currently logged-in user."""
    confirm = request.get_json(force=True).get("confirm", "")
    if confirm != current_user.email:
        return jsonify({"error": "Confirmation email does not match."}), 400

    errors = []

    # Delete files from disk
    import shutil
    for path in [get_upload_dir(), get_attach_dir()]:
        if os.path.isdir(path):
            try:
                shutil.rmtree(path)
            except Exception as e:
                errors.append(f"Failed to delete {path}: {e}")
            else:
                if os.path.isdir(path):
                    errors.append(f"Directory still exists after deletion: {path}")

    # Drop the user's MongoDB database
    try:
        _get_client().drop_database(get_db().name)
    except Exception as e:
        errors.append(f"Failed to drop database: {e}")

    if errors:
        return jsonify({"error": "Partial failure: " + "; ".join(errors)}), 500

    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Admin — Log Viewer  (admin emails only)
# ---------------------------------------------------------------------------

@app.route("/admin/logs")
def admin_logs():
    if not is_admin():
        return "Forbidden", 403
    log_files = sorted(
        [f for f in os.listdir(_LOG_DIR) if f.startswith("session_activity")],
        reverse=True,
    )
    selected = request.args.get("file", log_files[0] if log_files else "")
    lines = []
    if selected and selected in log_files:
        path = os.path.join(_LOG_DIR, selected)
        with open(path, encoding="utf-8", errors="replace") as fh:
            lines = fh.readlines()
        lines = [l.rstrip() for l in reversed(lines)]  # newest first
    return render_template("admin_logs.html",
                           log_files=log_files,
                           selected=selected,
                           lines=lines)


@app.route("/admin/logs/download")
def admin_logs_download():
    if not is_admin():
        return "Forbidden", 403
    filename = request.args.get("file", "")
    safe = os.path.basename(filename)
    path = os.path.join(_LOG_DIR, safe)
    if not safe.startswith("session_activity") or not os.path.exists(path):
        return "Not found", 404
    return send_file(path, as_attachment=True, download_name=safe)


# ---------------------------------------------------------------------------
# Chunked / Resumable Upload  (replaces single-shot /upload for large files)
# ---------------------------------------------------------------------------
# Flow:
#   1. POST /upload/init          → {upload_id, received_bytes}
#   2. PUT  /upload/chunk         → {received_bytes}   (repeat until done)
#   3. POST /upload/finalise      → {job_id, filename}  (triggers import)
# ---------------------------------------------------------------------------

@app.route("/upload/init", methods=["POST"])
def upload_init():
    """Start (or resume) a chunked upload session."""
    data       = request.get_json(force=True)
    filename   = data.get("filename", "")
    total_size = int(data.get("total_size", 0))

    if not filename.lower().endswith(".pst"):
        return jsonify({"error": "Only .pst files are supported"}), 400

    # Quota check
    if not AUTH_DISABLED and not is_admin():
        quota    = get_user_quota_bytes()
        used     = get_user_storage_bytes()
        if used + total_size > quota:
            plan     = PLANS.get(get_user_plan_name(), PLANS["free"])
            used_gb  = used / 1024 ** 3
            quota_gb = quota / 1024 ** 3
            return jsonify({
                "error": (f"Storage quota exceeded — you are using "
                          f"{used_gb:.1f} GB of your {quota_gb:.0f} GB "
                          f"{plan['name']} plan limit. "
                          f"Upgrade your plan to upload more."),
                "quota_exceeded": True,
                "upgrade_url":    "/billing",
            }), 413

    upload_dir = get_upload_dir()
    safe_name  = os.path.basename(filename)
    save_path  = os.path.join(upload_dir, safe_name)
    tmp_path   = save_path + ".part"

    # How many bytes already received (supports resume)
    received = os.path.getsize(tmp_path) if os.path.exists(tmp_path) else 0

    # Stable upload_id = user_dir + filename hash
    upload_id = str(uuid.uuid5(uuid.NAMESPACE_URL, tmp_path))

    return jsonify({"upload_id": upload_id, "received_bytes": received,
                    "tmp_path": tmp_path, "save_path": save_path})


@app.route("/upload/chunk", methods=["PUT", "POST"])
def upload_chunk():
    """Append a single chunk to the .part file."""
    tmp_path   = request.headers.get("X-Tmp-Path", "")
    offset     = int(request.headers.get("X-Offset", 0))

    if not tmp_path or ".." in tmp_path:
        return jsonify({"error": "Invalid path"}), 400

    chunk_data = request.get_data()

    with open(tmp_path, "ab") as fh:
        # If client is resuming from a known offset, seek is not needed for
        # append-only — but guard against duplicate chunks
        current_size = os.path.getsize(tmp_path) if os.path.exists(tmp_path) else 0
        if offset != current_size:
            # Client re-sent a chunk we already have — skip it
            return jsonify({"received_bytes": current_size})
        fh.write(chunk_data)

    received = os.path.getsize(tmp_path)
    return jsonify({"received_bytes": received})


@app.route("/upload/finalise", methods=["POST"])
def upload_finalise():
    """Rename .part → final file and kick off the import job."""
    data      = request.get_json(force=True)
    tmp_path  = data.get("tmp_path", "")
    save_path = data.get("save_path", "")

    if not tmp_path or not save_path or ".." in tmp_path or ".." in save_path:
        return jsonify({"error": "Invalid path"}), 400
    if not os.path.exists(tmp_path):
        return jsonify({"error": "Upload not found — please restart the upload"}), 404

    os.rename(tmp_path, save_path)

    # Capture per-user context
    attach_dir = get_attach_dir()
    user_db    = get_db().name
    filename   = os.path.basename(save_path)

    job_id = str(uuid.uuid4())
    q = queue.Queue()
    jobs[job_id] = {"queue": q, "status": "running", "filename": filename}
    user_email = current_user.email if current_user.is_authenticated else ""

    t = threading.Thread(target=_run_import,
                         args=(job_id, save_path, user_db, attach_dir, q, user_email), daemon=True)
    t.start()

    return jsonify({"job_id": job_id, "filename": filename})


@app.route("/upload", methods=["POST"])
def upload():
    """Legacy single-shot upload (kept for small files / backwards compat)."""
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400
    if not f.filename.lower().endswith(".pst"):
        return jsonify({"error": "Only .pst files are supported"}), 400

    # ── Quota check ────────────────────────────────────────────────────────────
    if not AUTH_DISABLED and not is_admin():
        quota   = get_user_quota_bytes()
        used    = get_user_storage_bytes()
        incoming = request.content_length or 0
        if used + incoming > quota:
            plan    = PLANS.get(get_user_plan_name(), PLANS["free"])
            used_gb = used / 1024 ** 3
            quota_gb= quota / 1024 ** 3
            return jsonify({
                "error": (f"Storage quota exceeded — you are using "
                          f"{used_gb:.1f} GB of your {quota_gb:.0f} GB "
                          f"{plan['name']} plan limit. "
                          f"Upgrade your plan to upload more."),
                "quota_exceeded": True,
                "upgrade_url": "/billing",
            }), 413

    # Capture per-user paths BEFORE the thread (no request context in thread)
    upload_dir  = get_upload_dir()
    attach_dir  = get_attach_dir()
    user_db     = get_db().name

    save_path = os.path.join(upload_dir, f.filename)
    f.save(save_path)

    job_id = str(uuid.uuid4())
    q = queue.Queue()
    jobs[job_id] = {"queue": q, "status": "running", "filename": f.filename}
    user_email = current_user.email if current_user.is_authenticated else ""

    t = threading.Thread(target=_run_import,
                         args=(job_id, save_path, user_db, attach_dir, q, user_email), daemon=True)
    t.start()

    return jsonify({"job_id": job_id, "filename": f.filename})


def _clamscan(path: str, q: queue.Queue) -> tuple[bool, str]:
    """
    Scan a file with ClamAV. Returns (clean, message).
    Emits keep-alive dots into q every 10 seconds so the browser stays connected.
    """
    # clamdscan runs as the clamav user — make file world-readable so it can access it
    try:
        orig_mode = os.stat(path).st_mode
        os.chmod(path, orig_mode | 0o004)
    except Exception:
        orig_mode = None

    def _restore_mode():
        if orig_mode is not None:
            try: os.chmod(path, orig_mode)
            except Exception: pass

    for scanner in ("clamscan", "clamdscan"):
        try:
            proc = subprocess.Popen(
                [scanner, "--no-summary", "--max-filesize=4000M", "--max-scansize=4000M", path],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                text=True, encoding="utf-8", errors="replace",
            )
            # Read stdout and stderr in threads so we can emit keep-alives
            output_lines: list = []
            stderr_lines: list = []
            done_event = threading.Event()

            def _read():
                for line in proc.stdout:
                    output_lines.append(line)
                done_event.set()

            def _read_err():
                for line in proc.stderr:
                    stderr_lines.append(line)

            threading.Thread(target=_read, daemon=True).start()
            threading.Thread(target=_read_err, daemon=True).start()

            elapsed = 0
            while not done_event.wait(timeout=10):
                elapsed += 10
                q.put(f"Scanning… {elapsed}s elapsed")

            proc.wait()
            stdout = "".join(output_lines).strip()
            stderr = "".join(stderr_lines).strip()
            combined = (stdout or stderr)[:200]
            _restore_mode()

            if proc.returncode == 0:
                return True, "Clean"
            if proc.returncode == 1:
                for line in stdout.splitlines():
                    if "FOUND" in line:
                        return False, line.strip()
                return False, "Virus detected"
            # returncode 2 = error
            return True, f"Scan error (skipped): {combined}"

        except FileNotFoundError:
            continue
        except Exception as e:
            _restore_mode()
            return True, f"Scan skipped: {e}"

    _restore_mode()
    return True, "ClamAV not installed — scan skipped"


def _run_import(job_id: str, pst_path: str, user_db: str, attach_dir: str,
                q: queue.Queue, user_email: str = ""):
    """Run pst_to_mongodb.py in a subprocess and feed stdout into the queue."""
    filename = os.path.basename(pst_path)

    # ── Virus scan before import ──────────────────────────────────────────────
    q.put(f"Scanning {filename} for viruses…")
    clean, scan_msg = _clamscan(pst_path, q)
    if not clean:
        q.put(f"ERROR: File rejected — {scan_msg}")
        jobs[job_id]["status"] = "error"
        try:
            os.remove(pst_path)
        except Exception:
            pass
        _send_notification_email(
            user_email,
            "PST upload rejected — virus detected",
            f"Your file '{filename}' was rejected: {scan_msg}",
            f"<p>Your file <strong>{filename}</strong> was rejected by our virus scanner.</p>"
            f"<p>Detection: <code>{scan_msg}</code></p>"
            f"<p>If you believe this is a false positive, please contact support.</p>",
        )
        q.put(None)
        return
    q.put(f"Virus scan passed: {scan_msg}")
    _send_notification_email(
        user_email,
        f"✅ Virus scan passed — {filename}",
        f"Your file '{filename}' passed the virus scan and is now being imported.",
        f"<p>Your file <strong>{filename}</strong> passed the virus scan successfully.</p>"
        f"<p>The import is now running. Large files can take several hours. You will receive an email after each step completes.</p>"
        f"<p><a href='https://pstbrowser.com/process-flow' style='background:#2563eb;color:#fff;"
        f"padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:600;"
        f"display:inline-block'>View processing status</a></p>",
    )

    cmd = [
        sys.executable, "pst_to_mongodb.py",
        "--pst",        pst_path,
        "--mongo",      MONGO_URI,
        "--db",         user_db,
        "--attach-dir", attach_dir,
    ]
    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            cwd=os.path.dirname(__file__),
        )
        for line in proc.stdout:
            q.put(line.rstrip())
        proc.wait()
        status = "done" if proc.returncode == 0 else "error"
        jobs[job_id]["status"] = status
        if status == "done":
            _send_notification_email(
                user_email,
                f"✅ PST import complete — {filename}",
                f"Your file '{filename}' has been imported successfully. Indexing will begin shortly.",
                f"<p>Your file <strong>{filename}</strong> has been imported successfully.</p>"
                f"<p>Search indexing is now starting. You will receive another email when the full-text index is ready.</p>"
                f"<p><a href='https://pstbrowser.com/process-flow' style='background:#2563eb;color:#fff;"
                f"padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:600;"
                f"display:inline-block'>View processing status</a></p>",
            )
    except Exception as e:
        q.put(f"ERROR: {e}")
        jobs[job_id]["status"] = "error"
    finally:
        q.put(None)  # sentinel — signals stream end


@app.route("/progress/<job_id>")
def progress(job_id: str):
    if job_id not in jobs:
        return jsonify({"error": "Job not found"}), 404

    def generate():
        q = jobs[job_id]["queue"]
        while True:
            try:
                line = q.get(timeout=30)
            except queue.Empty:
                yield "data: \n\n"   # keep-alive
                continue
            if line is None:
                status = jobs[job_id].get("status", "done")
                yield f"data: __DONE__{status}__\n\n"
                break
            yield f"data: {json.dumps(line)}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


def _build_index_progress(col, q: queue.Queue, user_email: str = ""):
    """Run in a background thread: build the text index and emit JSON progress events."""
    import time

    def emit(msg, pct, ok=True):
        q.put(json.dumps({"msg": msg, "pct": pct, "ok": ok}))

    emit("Counting documents…", 5)
    total = col.count_documents({})
    emit(f"Building search index on {total:,} documents…", 10)

    # Run _ensure_text_index in its own thread so we can animate while it blocks
    result: dict = {"done": False, "error": None}

    def _create():
        try:
            _ensure_text_index(col)
        except Exception as e:
            result["error"] = str(e)
        finally:
            result["done"] = True

    worker = threading.Thread(target=_create, daemon=True)
    worker.start()

    pct = 15
    while not result["done"]:
        time.sleep(0.6)
        pct = min(pct + 5, 90)
        emit(f"Building search index… {pct}%", pct)

    if result["error"]:
        emit(f"Index error: {result['error']}", 100, ok=False)
    else:
        emit("Search index ready!", 100, ok=True)
        _send_notification_email(
            user_email,
            "✅ Indexing complete — pstbrowser.com",
            "Your search index has finished building. Full-text search is now available at https://pstbrowser.com",
            "<p>Your search index has finished building. You can now search across all your emails by keyword, sender, date, and more.</p>"
            "<p>If you have PDF attachments, OCR processing may still be running to make document content searchable.</p>"
            "<p><a href='https://pstbrowser.com' style='background:#2563eb;color:#fff;"
            "padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:600;"
            "display:inline-block'>Search your emails</a></p>"
            "<p style='margin-top:8px'><a href='https://pstbrowser.com/process-flow' style='color:#2563eb;font-size:.9rem;'>View full processing status</a></p>",
        )

    q.put(None)   # sentinel


@app.route("/build-index")
def build_index_route():
    """SSE stream that builds the text index and reports progress."""
    col = get_col()
    q: queue.Queue = queue.Queue()
    user_email = current_user.email if current_user.is_authenticated else ""
    threading.Thread(target=_build_index_progress, args=(col, q, user_email), daemon=True).start()

    def generate():
        while True:
            try:
                item = q.get(timeout=120)
            except queue.Empty:
                yield "data: \n\n"   # keep-alive
                continue
            if item is None:
                break
            yield f"data: {item}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


def _ensure_text_index(col):
    """Create (or recreate) the text index covering subject, body and attachment text."""
    spec    = [("subject", "text"), ("body_plain", "text"), ("attachment_text", "text")]
    weights = {"subject": 10, "body_plain": 5, "attachment_text": 3}
    try:
        col.create_index(spec, name="text_search",
                         default_language="english", weights=weights)
    except Exception:
        # Index exists with a different definition — drop and recreate
        try:
            col.drop_index("text_search")
            col.create_index(spec, name="text_search",
                             default_language="english", weights=weights)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Atlas Full-Text Search (requires MongoDB Atlas or Atlas Local)
# Falls back to standard $text index when unavailable.
# ---------------------------------------------------------------------------
_ATLAS_INDEX = "atlas_fts"
_ATLAS_AVAIL: dict = {}   # {db_name: True|False} — per-user cache

_ATLAS_INDEX_DEF = {
    "mappings": {
        "dynamic": False,
        "fields": {
            "subject":         {"type": "string", "analyzer": "lucene.english"},
            "body_plain":      {"type": "string", "analyzer": "lucene.english"},
            "attachment_text": {"type": "string", "analyzer": "lucene.english"},
            "from_addr":       {"type": "string", "analyzer": "lucene.keyword"},
            "to_addrs":        {"type": "string", "analyzer": "lucene.keyword"},
            "folder_path":     {"type": "string", "analyzer": "lucene.standard"},
            "date":            {"type": "date"},
            "item_type":       {"type": "string", "analyzer": "lucene.keyword"},
            "has_attachments": {"type": "boolean"},
        },
    }
}


def _check_atlas(col) -> bool:
    """Probe whether Atlas Search index atlas_fts is usable (result cached per user DB)."""
    key = col.database.name
    if key in _ATLAS_AVAIL:
        return _ATLAS_AVAIL[key]
    try:
        list(col.aggregate([
            {"$search": {"index": _ATLAS_INDEX,
                         "text": {"query": "probe", "path": "subject"}}},
            {"$limit": 1},
        ]))
        _ATLAS_AVAIL[key] = True
    except Exception:
        _ATLAS_AVAIL[key] = False
    return _ATLAS_AVAIL[key]


def _create_atlas_index(col) -> bool:
    """Try to create (or confirm existence of) the atlas_fts search index."""
    key = col.database.name
    _ATLAS_AVAIL.pop(key, None)   # reset cache so next _check_atlas re-probes
    try:
        existing = list(col.list_search_indexes())
        if any(i.get("name") == _ATLAS_INDEX for i in existing):
            _ATLAS_AVAIL[key] = True
            return True
    except Exception:
        pass
    try:
        col.create_search_index({"name": _ATLAS_INDEX, "definition": _ATLAS_INDEX_DEF})
        _ATLAS_AVAIL[key] = True
        return True
    except Exception:
        _ATLAS_AVAIL[key] = False
        return False


def _build_atlas_operator(q: str, fuzzy: bool = False,
                           filter_clauses=None) -> dict:
    """
    Parse a query string into a MongoDB Atlas Search compound operator.

    Syntax supported:
      plain terms      – must appear anywhere in the indexed fields
      "quoted phrase"  – exact phrase required
      -excluded        – term must NOT appear
    """
    import shlex as _shlex

    phrases, terms, excluded = [], [], []
    try:
        for tok in _shlex.split(q):
            if " " in tok:
                phrases.append(tok)
            elif tok.startswith("-") and len(tok) > 1:
                excluded.append(tok[1:])
            else:
                terms.append(tok)
    except ValueError:
        terms = q.split()

    all_paths = ["subject", "body_plain", "attachment_text"]
    must, should, must_not = [], [], []

    if terms:
        plain = " ".join(terms)
        text_q: dict = {"query": plain, "path": all_paths}
        if fuzzy and len(plain) > 3:
            text_q["fuzzy"] = {"maxEdits": 1, "prefixLength": 2}
        must.append({"text": text_q})
        # Boost subject hits in relevance scoring
        should.append({"text": {
            "query": plain,
            "path":  "subject",
            "score": {"boost": {"value": 8}},
        }})

    for ph in phrases:
        must.append({"phrase": {"query": ph, "path": all_paths}})
        should.append({"phrase": {
            "query": ph,
            "path":  "subject",
            "score": {"boost": {"value": 12}},
        }})

    for ex in excluded:
        must_not.append({"text": {"query": ex, "path": all_paths}})

    if not must:
        must = [{"text": {"query": q, "path": all_paths}}]

    compound: dict = {"must": must}
    if should:
        compound["should"] = should
    if must_not:
        compound["mustNot"] = must_not
    if filter_clauses:
        compound["filter"] = filter_clauses

    return {"compound": compound}


def _highlights_to_html(highlights: list, q: str = "") -> str:
    """
    Convert Atlas Search highlight metadata into an HTML snippet.
    Hit spans are wrapped in <mark class="search-hit">.
    Prefers body_plain passages, falls back to subject.
    """
    import html as _h
    if not highlights:
        return ""
    best = next((h for h in highlights if h.get("path") == "body_plain"), None)
    if best is None:
        best = next((h for h in highlights if h.get("path") == "subject"), None)
    if best is None:
        best = highlights[0] if highlights else None
    if not best:
        return ""
    parts = []
    for seg in best.get("texts", []):
        val = _h.escape(seg.get("value", ""))
        if seg.get("type") == "hit":
            parts.append(f'<mark class="search-hit">{val}</mark>')
        else:
            parts.append(val)
    return "".join(parts)


def _plain_snippet(body: str, q: str = "", length: int = 200) -> str:
    """
    Extract a plain-text excerpt positioned near the first query-term occurrence.
    Used as a fallback when Atlas Search highlights are not available.
    """
    if not body:
        return ""
    flat = " ".join(body.split())
    pos = 0
    if q:
        flat_lower = flat.lower()
        for tok in q.lower().split():
            clean = tok.strip('"').strip("'")
            if len(clean) > 2:
                p = flat_lower.find(clean)
                if p >= 0:
                    pos = max(0, p - 60)
                    break
    raw = flat[pos: pos + length]
    if pos > 0:
        raw = "…" + raw
    if pos + length < len(flat):
        raw += "…"
    return raw


def _run_atlas_search(col, q: str, item_type: str, date_filter: dict,
                       has_att: str, folder_q: str, from_addr_q: str,
                       to_addr_q: str, skip: int, limit: int,
                       fuzzy: bool = False) -> dict:
    """
    Execute an Atlas Full-Text Search query.
    Returns {"rows": [...], "total": int, "mode": "atlas"}.
    Raises on failure so the caller can fall back to $text search.
    """
    # ── Atlas-native filter clauses (equality / range — no regex) ────────────
    atlas_filters: list = []
    if item_type and item_type not in ("all", "hasatt"):
        atlas_filters.append({"text": {"query": item_type, "path": "item_type"}})
    if item_type == "hasatt" or has_att == "1":
        atlas_filters.append({"equals": {"path": "has_attachments", "value": True}})
    elif has_att == "0":
        atlas_filters.append({"equals": {"path": "has_attachments", "value": False}})
    dr: dict = {"path": "date"}
    if date_filter.get("$gte"):
        dr["gte"] = date_filter["$gte"]
    if date_filter.get("$lte"):
        dr["lte"] = date_filter["$lte"]
    if len(dr) > 1:
        atlas_filters.append({"range": dr})

    operator  = _build_atlas_operator(q, fuzzy=fuzzy,
                                       filter_clauses=atlas_filters or None)
    search_stage = {
        "$search": {
            "index": _ATLAS_INDEX,
            **operator,
            "highlight": {
                "path":              ["subject", "body_plain"],
                "maxCharsToExamine": 3000,
                "maxNumPassages":    2,
            },
        }
    }

    # ── Post-search regex filters (can't live inside $search) ─────────────────
    post_match: dict = {}
    if from_addr_q:
        post_match["from_addr"] = {"$regex": from_addr_q, "$options": "i"}
    if to_addr_q:
        rx = {"$regex": to_addr_q, "$options": "i"}
        post_match["$or"] = [{"to_addrs": rx}, {"cc_addrs": rx}]
    if folder_q:
        post_match["folder_path"] = {"$regex": folder_q, "$options": "i"}

    base_pipe: list = [search_stage]
    if post_match:
        base_pipe.append({"$match": post_match})

    # ── Total count ───────────────────────────────────────────────────────────
    cr    = list(col.aggregate(base_pipe + [{"$count": "n"}], allowDiskUse=True))
    total = cr[0]["n"] if cr else 0

    # ── Fetch documents with highlights ───────────────────────────────────────
    doc_pipe = base_pipe + [
        {"$addFields": {
            "_score":      {"$meta": "searchScore"},
            "_highlights": {"$meta": "searchHighlights"},
        }},
        {"$skip": skip},
        {"$limit": limit},
        {"$project": {
            "subject": 1, "from_addr": 1, "date": 1,
            "has_attachments": 1, "item_type": 1,
            "body_plain": 1, "folder_path": 1, "to_addrs": 1,
            "tags": 1, "_score": 1, "_highlights": 1,
        }},
    ]
    docs = list(col.aggregate(doc_pipe, allowDiskUse=True))

    # ── Build row dicts ───────────────────────────────────────────────────────
    rows = []
    for doc in docs:
        body          = (doc.get("body_plain") or "").strip()
        highlights    = doc.get("_highlights") or []
        highlight_html = _highlights_to_html(highlights, q)
        snippet       = _plain_snippet(body, q) if body and not highlight_html else ""
        rows.append({
            "_id":             str(doc["_id"]),
            "subject":         doc.get("subject") or "(no subject)",
            "from_addr":       doc.get("from_addr") or "",
            "to_addrs":        doc.get("to_addrs") or [],
            "date":            doc["date"].strftime("%Y-%m-%d %H:%M") if doc.get("date") else "",
            "has_attachments": doc.get("has_attachments", False),
            "item_type":       doc.get("item_type") or "email",
            "folder_path":     doc.get("folder_path") or "",
            "snippet":         snippet,
            "highlight_html":  highlight_html,
            "score":           round(doc.get("_score", 0), 3),
            "tags":            doc.get("tags") or [],
        })

    _ATLAS_AVAIL[col.database.name] = True
    return {"rows": rows, "total": total, "mode": "atlas"}


@app.route("/email/<email_id>")
def get_email(email_id: str):
    col = get_col()
    doc = col.find_one({"_id": email_id})
    if not doc:
        return jsonify({"error": "Not found"}), 404
    doc["_id"] = str(doc["_id"])
    for field in ("date", "delivery_time", "creation_time", "modification_time",
                  "client_submit_time", "imported_at"):
        if isinstance(doc.get(field), datetime.datetime):
            doc[field] = doc[field].strftime("%Y-%m-%d %H:%M:%S")
    return jsonify(doc)


def _build_eml(doc: dict) -> bytes:
    """
    Build a RFC 2822 .eml byte-string from a PST record document.
    Shared by the single-email download and the bulk ZIP export.
    """
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    from email.header         import Header
    from email.utils          import formatdate
    import time as _time

    body_plain = (doc.get("body_plain") or "").strip()
    body_html  = (doc.get("body_html")  or "").strip()

    if body_plain and body_html:
        msg = MIMEMultipart("alternative")
        msg.attach(MIMEText(body_plain, "plain", "utf-8"))
        msg.attach(MIMEText(body_html,  "html",  "utf-8"))
    elif body_html:
        msg = MIMEText(body_html, "html", "utf-8")
    else:
        msg = MIMEText(body_plain or "", "plain", "utf-8")

    subject = doc.get("subject") or ""
    msg["Subject"] = Header(subject, "utf-8")
    msg["From"]    = doc.get("from_addr") or ""

    to_list = doc.get("to_addrs") or []
    if to_list:
        msg["To"] = ", ".join(to_list)
    cc_list = doc.get("cc_addrs") or []
    if cc_list:
        msg["CC"] = ", ".join(cc_list)

    date_val = doc.get("date")
    if isinstance(date_val, datetime.datetime):
        msg["Date"] = formatdate(_time.mktime(date_val.timetuple()), localtime=True)
    elif date_val:
        msg["Date"] = str(date_val)

    msg_id = doc.get("message_id")
    if msg_id:
        msg["Message-ID"] = msg_id

    msg["X-PST-Folder"] = doc.get("folder_path") or ""
    return msg.as_bytes()


def _eml_filename(doc: dict, prefix: str = "") -> str:
    """Return a safe .eml filename derived from the email subject."""
    import re as _re
    subject = doc.get("subject") or ""
    slug = _re.sub(r"[^\w\s-]", "", subject)[:60].strip()
    slug = _re.sub(r"\s+", "_", slug) or "email"
    return (prefix + slug + ".eml") if prefix else (slug + ".eml")


@app.route("/email/<email_id>/download/eml")
def download_email_eml(email_id: str):
    """Stream a single RFC 2822 .eml file for one email record."""
    col = get_col()
    doc = col.find_one({"_id": email_id})
    if not doc:
        return "Not found", 404
    return Response(
        _build_eml(doc),
        mimetype="message/rfc822",
        headers={"Content-Disposition": f'attachment; filename="{_eml_filename(doc)}"'},
    )


@app.route("/email/<email_id>/print")
def print_email(email_id: str):
    """
    Render a print-optimised HTML page for a single email.
    Opens in a new tab; the user saves it as PDF via the browser's print dialog.
    """
    import html as _html

    col = get_col()
    doc = col.find_one({"_id": email_id})
    if not doc:
        return "Email not found", 404

    subject    = doc.get("subject") or "(no subject)"
    from_addr  = doc.get("from_addr") or ""
    to_addrs   = ", ".join(doc.get("to_addrs")  or [])
    cc_addrs   = ", ".join(doc.get("cc_addrs")  or [])
    folder     = doc.get("folder_path") or ""
    body_html  = (doc.get("body_html")  or "").strip()
    body_plain = (doc.get("body_plain") or "").strip()

    date_val = doc.get("date")
    if isinstance(date_val, datetime.datetime):
        date_str = date_val.strftime("%A, %d %B %Y, %H:%M")
    else:
        date_str = str(date_val or "")

    # Build meta table rows
    meta_rows = ""
    for label, val in [
        ("From",    from_addr),
        ("To",      to_addrs),
        ("CC",      cc_addrs),
        ("Date",    date_str),
        ("Folder",  folder),
    ]:
        if val:
            meta_rows += (
                f"<tr>"
                f"<th>{_html.escape(label)}</th>"
                f"<td>{_html.escape(val)}</td>"
                f"</tr>\n"
            )

    # Body content
    if body_html:
        body_section = f'<div class="email-body">{body_html}</div>'
    elif body_plain:
        body_section = (
            '<pre class="email-body-plain">'
            + _html.escape(body_plain)
            + "</pre>"
        )
    else:
        body_section = '<p class="no-body">No body content.</p>'

    page = (
        "<!DOCTYPE html>"
        "<html lang='en'>"
        "<head>"
        "<meta charset='UTF-8'>"
        f"<title>{_html.escape(subject)}</title>"
        "<style>"
        "*, *::before, *::after { box-sizing: border-box; }"
        "body { font-family: 'Segoe UI', Arial, sans-serif; font-size: 11pt;"
        "       color: #1a1a1a; margin: 0; padding: 16mm 20mm; }"
        "@media print {"
        "  body { padding: 0; }"
        "  @page { margin: 16mm 20mm; }"
        "  .no-print { display: none !important; }"
        "  .email-body img { max-width: 100% !important; }"
        "}"
        "h1.subject { font-size: 15pt; font-weight: 700; margin: 0 0 12px 0;"
        "             color: #1a1f36; line-height: 1.35; word-break: break-word; }"
        "table.meta { border-collapse: collapse; margin-bottom: 14px; font-size: 9.5pt; }"
        "table.meta th { color: #6b7280; font-weight: 600; text-align: right;"
        "                padding: 2px 10px 2px 0; white-space: nowrap; vertical-align: top; }"
        "table.meta td { color: #374151; padding: 2px 0; word-break: break-word; }"
        "hr.divider { border: none; border-top: 2px solid #1a1f36; margin: 14px 0 18px 0; }"
        ".email-body { font-size: 10.5pt; line-height: 1.65; }"
        ".email-body img { max-width: 100%; height: auto; }"
        ".email-body pre, .email-body code { white-space: pre-wrap; word-break: break-word;"
        "  font-size: .85em; background: #f4f6f9; padding: 2px 4px; }"
        ".email-body table { border-collapse: collapse; max-width: 100%; }"
        ".email-body td, .email-body th { border: 1px solid #e5e7eb; padding: 3px 7px; }"
        ".email-body-plain { font-family: 'Consolas', monospace; font-size: 9.5pt;"
        "  white-space: pre-wrap; word-break: break-word; line-height: 1.5;"
        "  background: #f9fafb; padding: 14px 16px; border-radius: 6px; margin: 0; }"
        ".no-body { color: #9ca3af; font-style: italic; }"
        ".print-bar { display: flex; align-items: center; gap: 10px; padding: 10px 16px;"
        "  background: #1a1f36; color: #fff; margin: -16mm -20mm 18px -20mm; }"
        "@media print { .print-bar { display: none; } }"
        ".print-bar span { flex: 1; font-size: 12px; opacity: .75; }"
        ".print-bar button { background: #fff; color: #1a1f36; border: none; border-radius: 5px;"
        "  padding: 6px 16px; font-size: 12px; font-weight: 600; cursor: pointer; }"
        ".print-bar button:hover { background: #e0e7ff; }"
        "</style>"
        "</head>"
        "<body>"
        "<div class='print-bar no-print'>"
        f"  <span>{_html.escape(subject)}</span>"
        "  <button onclick='window.print()'>&#128438; Save as PDF / Print</button>"
        "  <button onclick='window.close()' style='background:transparent;color:#fff;border:1px solid rgba(255,255,255,.4)'>Close</button>"
        "</div>"
        f"<h1 class='subject'>{_html.escape(subject)}</h1>"
        f"<table class='meta'>{meta_rows}</table>"
        "<hr class='divider'>"
        f"{body_section}"
        "<script>if(document.referrer||window.opener){{setTimeout(()=>window.print(),400);}}</script>"
        "</body>"
        "</html>"
    )

    return Response(page, mimetype="text/html; charset=utf-8")


@app.route("/retag", methods=["POST"])
def retag():
    """
    Tag existing documents with item_type based on folder_path keywords.
    Safe to run multiple times — only updates docs that have no item_type yet,
    unless force=true is passed.
    """
    col   = get_col()
    force = request.json.get("force", False) if request.is_json else False
    base  = {} if force else {"item_type": {"$exists": False}}

    def tag(extra_filter, type_val):
        q = {**base, **extra_filter}
        return col.update_many(q, {"$set": {"item_type": type_val}}).modified_count

    contacts     = tag({"folder_path": {"$regex": "contact",              "$options": "i"}}, "contact")
    appointments = tag({"folder_path": {"$regex": "calendar|appointment", "$options": "i"}}, "appointment")
    tasks        = tag({"folder_path": {"$regex": "task",                 "$options": "i"}}, "task")
    # Everything else → email
    emails       = tag({"item_type": {"$exists": False}}, "email")

    col.create_index("item_type")
    return jsonify({
        "emails": emails, "contacts": contacts,
        "appointments": appointments, "tasks": tasks
    })


@app.route("/api/search-mode")
def api_search_mode():
    """Return the active search mode: atlas or text."""
    col  = get_col()
    avail = _check_atlas(col)
    return jsonify({
        "mode":  "atlas" if avail else "text",
        "index": _ATLAS_INDEX,
    })


@app.route("/api/atlas-setup", methods=["POST"])
def api_atlas_setup():
    """Try to create the Atlas Search index (requires MongoDB Atlas)."""
    col = get_col()
    ok  = _create_atlas_index(col)
    return jsonify({
        "success": ok,
        "mode":    "atlas" if ok else "text",
        "message": (
            "Atlas Search index created. It may take 1–2 minutes to build before "
            "searches switch to Atlas mode."
            if ok else
            "Could not create index. Are you running MongoDB Atlas or Atlas Local?"
        ),
    })


@app.route("/type_counts")
def type_counts():
    """Return per-type document counts."""
    col      = get_col()
    pipeline = [{"$group": {"_id": "$item_type", "count": {"$sum": 1}}}]
    result   = {r["_id"] or "email": r["count"] for r in col.aggregate(pipeline)}
    total    = sum(result.values())
    result["all"]    = total
    result["hasatt"] = col.count_documents({"has_attachments": True})
    return jsonify(result)


@app.route("/records")
def records():
    """
    Unified browse + search endpoint.
    Supports: q, date_from, date_to, type, from_addr, to_addr, has_att,
              folder_path, page, per_page, sort, order, fuzzy.
    Tries Atlas Search when a query is present; falls back to $text index.
    """
    page      = max(1, int(request.args.get("page", 1)))
    per_page  = min(int(request.args.get("per_page", 50)), 200)
    sort_by   = request.args.get("sort", "date")
    order     = int(request.args.get("order", -1))
    item_type = request.args.get("type", "all").strip().lower()
    q         = request.args.get("q",          "").strip()
    date_from = request.args.get("date_from",  "").strip()
    date_to   = request.args.get("date_to",    "").strip()
    from_addr_q = request.args.get("from_addr",   "").strip()
    to_addr_q   = request.args.get("to_addr",     "").strip()
    has_att     = request.args.get("has_att",     "").strip()
    folder_q      = request.args.get("folder_path", "").strip()
    folder_exact  = request.args.get("folder_exact", "0") == "1"
    tag_q         = request.args.get("tag",         "").strip()
    fuzzy         = request.args.get("fuzzy", "0") == "1"

    valid_sorts = {"date", "subject", "from_addr"}
    if sort_by not in valid_sorts:
        sort_by = "date"
    if order not in (-1, 1):
        order = -1

    col  = get_col()
    skip = (page - 1) * per_page

    # Build date filter (shared by Atlas and $text paths)
    date_filter: dict = {}
    if date_from:
        try:
            date_filter["$gte"] = datetime.datetime.strptime(date_from, "%Y-%m-%d")
        except ValueError:
            pass
    if date_to:
        try:
            date_filter["$lte"] = datetime.datetime.strptime(date_to, "%Y-%m-%d") \
                                   + datetime.timedelta(days=1)
        except ValueError:
            pass

    # ── Atlas Search path ────────────────────────────────────────────────────
    if q and _check_atlas(col):
        try:
            result = _run_atlas_search(
                col, q, item_type, date_filter, has_att,
                folder_q, from_addr_q, to_addr_q,
                skip=skip, limit=per_page, fuzzy=fuzzy,
            )
            return jsonify({
                "rows":        result["rows"],
                "total":       result["total"],
                "page":        page,
                "per_page":    per_page,
                "pages":       max(1, -(-result["total"] // per_page)),
                "sort":        "score",
                "order":       -1,
                "type":        item_type,
                "q":           q,
                "search_mode": "atlas",
                "fuzzy":       fuzzy,
            })
        except Exception:
            pass  # fall through to $text

    # ── Standard $text / browse path ─────────────────────────────────────────
    query: dict = {}
    if item_type and item_type != "all":
        if item_type == "hasatt":
            query["has_attachments"] = True
        else:
            query["item_type"] = item_type
    if q:
        _ensure_text_index(col)
        query["$text"] = {"$search": q}
    if date_filter:
        query["date"] = date_filter
    if from_addr_q:
        query["from_addr"] = {"$regex": from_addr_q, "$options": "i"}
    if to_addr_q:
        rx = {"$regex": to_addr_q, "$options": "i"}
        query["$or"] = [{"to_addrs": rx}, {"cc_addrs": rx}]
    if has_att == "1":
        query["has_attachments"] = True
    elif has_att == "0":
        query["has_attachments"] = {"$ne": True}
    if folder_q:
        if folder_exact:
            import re as _re
            # Match exactly this folder path OR any sub-path (trailing / or \)
            _escaped = _re.escape(folder_q.replace("\\", "/"))
            query["folder_path"] = {"$regex": _escaped + r"($|[/\\])", "$options": "i"}
        else:
            query["folder_path"] = {"$regex": folder_q, "$options": "i"}
    if tag_q:
        query["tags"] = tag_q

    total = col.count_documents(query)

    projection = {"subject": 1, "from_addr": 1, "date": 1,
                  "has_attachments": 1, "item_type": 1, "body_plain": 1,
                  "folder_path": 1, "to_addrs": 1, "tags": 1}
    if q:
        projection["score"] = {"$meta": "textScore"}
        sort_order = [("score", {"$meta": "textScore"})]
    else:
        sort_order = [(sort_by, order)]

    cursor = col.find(query, projection).sort(sort_order).skip(skip).limit(per_page)

    rows = []
    for doc in cursor:
        body    = (doc.get("body_plain") or "").strip()
        snippet = _plain_snippet(body, q) if q and body else (" ".join(body.split())[:200] if body else "")
        rows.append({
            "_id":             str(doc["_id"]),
            "subject":         doc.get("subject") or "(no subject)",
            "from_addr":       doc.get("from_addr") or "",
            "to_addrs":        doc.get("to_addrs") or [],
            "date":            doc["date"].strftime("%Y-%m-%d %H:%M") if doc.get("date") else "",
            "has_attachments": doc.get("has_attachments", False),
            "item_type":       doc.get("item_type") or "email",
            "folder_path":     doc.get("folder_path") or "",
            "snippet":         snippet,
            "highlight_html":  "",
            "score":           round(doc.get("score", 0), 3) if q else 0,
            "tags":            doc.get("tags") or [],
        })

    return jsonify({
        "rows":        rows,
        "total":       total,
        "page":        page,
        "per_page":    per_page,
        "pages":       max(1, -(-total // per_page)),
        "sort":        sort_by,
        "order":       order,
        "type":        item_type,
        "q":           q,
        "search_mode": "text" if q else "",
        "fuzzy":       False,
    })


# ---------------------------------------------------------------------------
# Thread view helpers
# ---------------------------------------------------------------------------
import re as _re
_REPLY_RE = _re.compile(r'^(?:re|fw|fwd|aw|wg|sv|vs|tr)\s*:\s*', _re.IGNORECASE)

def _normalize_subject(s: str) -> str:
    """Strip Re:/Fwd: etc. prefixes (up to 5 levels) and return lower-cased result."""
    s = (s or "").strip()
    for _ in range(5):
        ns = _REPLY_RE.sub("", s).strip()
        if ns == s:
            break
        s = ns
    return s.lower()


@app.route("/threads")
def threads():
    """
    Return email conversations grouped by normalised subject (Python-side).
    Params: q (subject filter), min_count (int, default 2), page, per_page.
    """
    page      = max(1, int(request.args.get("page",      1)))
    per_page  = min(int(request.args.get("per_page",    25)), 100)
    min_count = max(1, int(request.args.get("min_count", 2)))
    q         = request.args.get("q", "").strip()
    q_lower   = q.lower() if q else ""

    col = get_col()

    # Group by lowercased subject in MongoDB (fast, no regex ops needed)
    match: dict = {"item_type": "email"}
    if q:
        match["subject"] = {"$regex": q, "$options": "i"}   # narrow early

    pipeline = [
        {"$match": match},
        {"$group": {
            "_id":             {"$toLower": {"$ifNull": ["$subject", ""]}},
            "count":           {"$sum": 1},
            "latest_date":     {"$max": "$date"},
            "earliest_date":   {"$min": "$date"},
            "senders":         {"$addToSet": "$from_addr"},
            "latest_subject":  {"$last":  "$subject"},
            "has_attachments": {"$max": {"$cond": [{"$eq": ["$has_attachments", True]}, 1, 0]}},
        }},
        {"$sort": {"latest_date": -1}},
        {"$limit": 10_000},   # cap for in-memory merge
    ]

    try:
        raw_rows = list(col.aggregate(pipeline, allowDiskUse=True))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # Merge groups that share the same normalised key (Re:/Fwd: stripped)
    merged: dict = {}
    for row in raw_rows:
        key = _normalize_subject(row["_id"] or "")
        if key not in merged:
            merged[key] = {
                "thread_key":     key,
                "count":          0,
                "latest_date":    None,
                "earliest_date":  None,
                "senders":        set(),
                "subject":        "",
                "has_attachments":False,
            }
        m = merged[key]
        m["count"] += row["count"]
        ld = row.get("latest_date")
        ed = row.get("earliest_date")
        if ld and (m["latest_date"] is None or ld > m["latest_date"]):
            m["latest_date"] = ld
            m["subject"]     = row.get("latest_subject") or ""
        if ed and (m["earliest_date"] is None or ed < m["earliest_date"]):
            m["earliest_date"] = ed
        m["senders"].update(s for s in (row.get("senders") or []) if s)
        m["has_attachments"] = m["has_attachments"] or bool(row.get("has_attachments"))

    # Filter by min_count and optional subject keyword
    thread_list = [
        m for m in merged.values()
        if m["count"] >= min_count and (not q_lower or q_lower in m["thread_key"])
    ]
    thread_list.sort(key=lambda x: x["latest_date"] or datetime.datetime.min, reverse=True)

    total    = len(thread_list)
    start    = (page - 1) * per_page
    page_rows = thread_list[start: start + per_page]

    result = []
    for m in page_rows:
        result.append({
            "thread_key":     m["thread_key"],
            "count":          m["count"],
            "latest_date":    m["latest_date"].strftime("%Y-%m-%d")   if m["latest_date"]   else "",
            "earliest_date":  m["earliest_date"].strftime("%Y-%m-%d") if m["earliest_date"] else "",
            "senders":        sorted(m["senders"])[:5],
            "subject":        m["subject"] or "(no subject)",
            "has_attachments":m["has_attachments"],
        })

    return jsonify({
        "threads":  result,
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "pages":    max(1, -(-total // per_page)),
        "q":        q,
    })


@app.route("/thread-emails")
def thread_emails():
    """Return all emails whose normalised subject matches the thread key, oldest first."""
    thread_key = request.args.get("key", "").strip()
    if not thread_key:
        return jsonify({"error": "key is required"}), 400

    col = get_col()

    # Use thread_key as a regex pre-filter, then verify exact normalised match in Python
    try:
        escaped = _re.escape(thread_key)
        cursor  = col.find(
            {"item_type": "email",
             "subject":   {"$regex": escaped, "$options": "i"}},
            {"from_addr": 1, "to_addrs": 1, "date": 1, "body_plain": 1,
             "has_attachments": 1, "subject": 1, "folder_path": 1}
        ).sort("date", 1).limit(2000)
        docs = list(cursor)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    emails = []
    for doc in docs:
        if _normalize_subject(doc.get("subject", "")) != thread_key:
            continue   # not the same thread
        body = (doc.get("body_plain") or "").strip()
        emails.append({
            "_id":             str(doc["_id"]),
            "subject":         doc.get("subject") or "(no subject)",
            "from_addr":       doc.get("from_addr") or "",
            "to_addrs":        doc.get("to_addrs") or [],
            "date":            doc["date"].strftime("%Y-%m-%d %H:%M") if doc.get("date") else "",
            "has_attachments": doc.get("has_attachments", False),
            "folder_path":     doc.get("folder_path") or "",
            "snippet":         " ".join(body.split())[:300] if body else "",
        })

    return jsonify({"thread_key": thread_key, "emails": emails, "count": len(emails)})


@app.route("/address-book")
def address_book():
    """
    Extract every unique email address seen in from_addr / to_addrs / cc_addrs,
    parse display names, count appearances, and return a searchable/sortable list.
    Params: q, sort (count|name|email|last_seen), page, per_page.
    """
    from email.utils import parseaddr as _parseaddr

    page     = max(1, int(request.args.get("page",      1)))
    per_page = min(int(request.args.get("per_page",    50)), 200)
    q        = request.args.get("q",    "").strip().lower()
    sort_by  = request.args.get("sort", "count")

    col = get_col()

    # Pull all address strings from every document in one pass
    pipeline = [
        {"$project": {
            "date": 1,
            "addresses": {"$concatArrays": [
                [{"$ifNull": ["$from_addr", ""]}],
                {"$ifNull": ["$to_addrs",   []]},
                {"$ifNull": ["$cc_addrs",   []]},
            ]},
        }},
        {"$unwind": "$addresses"},
        {"$match":  {"addresses": {"$nin": [None, ""]}}},
        {"$group": {
            "_id":        "$addresses",
            "count":      {"$sum": 1},
            "last_seen":  {"$max": "$date"},
            "first_seen": {"$min": "$date"},
        }},
        {"$sort":  {"count": -1}},
        {"$limit": 100_000},
    ]

    try:
        raw_rows = list(col.aggregate(pipeline, allowDiskUse=True))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # Parse name + address, merge case-insensitively
    merged: dict = {}
    for row in raw_rows:
        raw = (row["_id"] or "").strip()
        if not raw:
            continue
        name, addr = _parseaddr(raw)
        # Fallback: treat bare token as address if it looks like one
        if not addr or "@" not in addr:
            if "@" in raw and " " not in raw.strip("<> "):
                addr = raw.strip("<> ")
                name = ""
            else:
                continue
        addr_key = addr.lower()
        if addr_key not in merged:
            merged[addr_key] = {
                "email":      addr_key,
                "name":       (name or "").strip(),
                "count":      0,
                "last_seen":  None,
                "first_seen": None,
            }
        m = merged[addr_key]
        m["count"] += row["count"]
        if not m["name"] and name:
            m["name"] = name.strip()
        ld, fd = row.get("last_seen"), row.get("first_seen")
        if ld and (m["last_seen"]  is None or ld > m["last_seen"]):  m["last_seen"]  = ld
        if fd and (m["first_seen"] is None or fd < m["first_seen"]): m["first_seen"] = fd

    # Filter
    entries = list(merged.values())
    if q:
        entries = [e for e in entries
                   if q in e["email"] or q in e["name"].lower()]

    # Sort
    if   sort_by == "name":      entries.sort(key=lambda x: (x["name"].lower() or x["email"], x["email"]))
    elif sort_by == "email":     entries.sort(key=lambda x: x["email"])
    elif sort_by == "last_seen": entries.sort(key=lambda x: x["last_seen"] or datetime.datetime.min, reverse=True)
    else:                        entries.sort(key=lambda x: x["count"], reverse=True)

    total     = len(entries)
    start     = (page - 1) * per_page
    page_data = entries[start: start + per_page]

    result = [{
        "email":      e["email"],
        "name":       e["name"],
        "count":      e["count"],
        "last_seen":  e["last_seen"].strftime("%Y-%m-%d")  if e["last_seen"]  else "",
        "first_seen": e["first_seen"].strftime("%Y-%m-%d") if e["first_seen"] else "",
    } for e in page_data]

    return jsonify({
        "addresses": result,
        "total":     total,
        "page":      page,
        "per_page":  per_page,
        "pages":     max(1, -(-total // per_page)),
        "q":         q,
        "sort":      sort_by,
    })


@app.route("/contact-card")
def contact_card():
    """
    Return rich contact data for a given email address.
    Merges stored contact records (item_type=contact) with address-book activity stats.
    """
    import re as _re
    from email.utils import parseaddr as _parseaddr

    email_q = request.args.get("email", "").strip().lower()
    if not email_q:
        return jsonify({"error": "email param required"}), 400

    col = get_col()

    # 1. Look up the best matching stored contact
    pattern = _re.compile(_re.escape(email_q), _re.IGNORECASE)
    contact_doc = col.find_one({
        "item_type": "contact",
        "$or": [
            {"from_addr":       {"$regex": pattern.pattern, "$options": "i"}},
            {"email_addresses": {"$elemMatch": {"$regex": pattern.pattern, "$options": "i"}}},
        ]
    })

    # 2. Address-book activity stats (count + date range)
    pipe_stats = [
        {"$project": {
            "date": 1,
            "addresses": {"$concatArrays": [
                [{"$ifNull": ["$from_addr", ""]}],
                {"$ifNull": ["$to_addrs", []]},
                {"$ifNull": ["$cc_addrs", []]},
            ]},
        }},
        {"$unwind": "$addresses"},
        {"$match":  {"addresses": {"$regex": pattern.pattern, "$options": "i"}}},
        {"$group": {
            "_id":        None,
            "count":      {"$sum": 1},
            "last_seen":  {"$max": "$date"},
            "first_seen": {"$min": "$date"},
        }},
    ]
    stats_rows = list(col.aggregate(pipe_stats, allowDiskUse=True))
    stats = stats_rows[0] if stats_rows else {}

    result = {
        "email":      email_q,
        "count":      stats.get("count", 0),
        "last_seen":  stats["last_seen"].strftime("%Y-%m-%d")  if stats.get("last_seen")  else "",
        "first_seen": stats["first_seen"].strftime("%Y-%m-%d") if stats.get("first_seen") else "",
        "has_stored_contact": bool(contact_doc),
    }

    if contact_doc:
        # Merge in every rich contact field
        for field in (
            "display_name", "first_name", "last_name", "middle_name",
            "honorific", "suffix", "nickname",
            "company_name", "department", "job_title", "office_location",
            "manager", "assistant", "spouse", "children",
            "birthday", "anniversary", "web_page", "im_address",
            "categories", "notes", "body_plain", "folder_path",
            "business_street", "business_street2",
            "business_city", "business_state", "business_postal", "business_country",
            "home_street", "home_city", "home_state", "home_postal", "home_country",
            "other_street", "other_city", "other_state", "other_postal", "other_country",
        ):
            val = contact_doc.get(field)
            if val:
                result[field] = val

        result["email_addresses"] = (
            contact_doc.get("email_addresses")
            or ([contact_doc.get("from_addr")] if contact_doc.get("from_addr") else [email_q])
        )
        result["phone_numbers"] = contact_doc.get("phone_numbers") or {}
    else:
        result["email_addresses"] = [email_q]
        result["phone_numbers"]   = {}

    return jsonify(result)


# ---------------------------------------------------------------------------
# Contact Duplicate Detection & Merge
# ---------------------------------------------------------------------------

@app.route("/contacts/duplicates")
def contact_duplicates():
    """
    Find contact records (item_type=contact) that share an email address
    or a normalised display name.
    Params: strategy (both|email|name), page, per_page.
    """
    col      = get_col()
    page     = max(1, int(request.args.get("page",     1)))
    per_page = min(int(request.args.get("per_page",   50)), 200)
    strategy = request.args.get("strategy", "both").strip().lower()

    groups       = []
    seen_id_sets = set()   # dedupe: same pair of IDs appearing under multiple keys

    # ── Strategy A: shared email address ────────────────────────────────────
    if strategy in ("both", "email"):
        pipe_a = [
            {"$match": {"item_type": "contact"}},
            {"$addFields": {
                "all_emails": {"$concatArrays": [
                    {"$cond": {
                        "if":   {"$and": [{"$gt": ["$from_addr", None]}, {"$gt": ["$from_addr", ""]}]},
                        "then": [{"$toLower": "$from_addr"}],
                        "else": [],
                    }},
                    {"$map": {
                        "input": {"$ifNull": ["$email_addresses", []]},
                        "as":    "e",
                        "in":    {"$toLower": "$$e"},
                    }},
                ]},
            }},
            {"$unwind": "$all_emails"},
            {"$match":  {"all_emails": {"$nin": [None, ""]}}},
            {"$group": {
                "_id":          "$all_emails",
                "count":        {"$sum": 1},
                "record_ids":   {"$push": "$_id"},
                "display_names": {"$addToSet": "$display_name"},
                "folder_paths": {"$addToSet": "$folder_path"},
            }},
            {"$match": {"count": {"$gt": 1}}},
            {"$sort":  {"count": -1}},
            {"$limit": 5_000},
        ]
        for row in col.aggregate(pipe_a, allowDiskUse=True):
            ids_key = frozenset(str(i) for i in row["record_ids"])
            if ids_key in seen_id_sets:
                continue
            seen_id_sets.add(ids_key)
            groups.append({
                "key":           row["_id"],
                "method":        "email",
                "count":         row["count"],
                "record_ids":    [str(i) for i in row["record_ids"]],
                "display_names": [n for n in (row.get("display_names") or []) if n],
                "folder_paths":  [f for f in (row.get("folder_paths")  or []) if f],
                "duplicates":    row["count"] - 1,
            })

    # ── Strategy B: shared normalised display name ────────────────────────────
    if strategy in ("both", "name"):
        pipe_b = [
            {"$match": {
                "item_type":    "contact",
                "display_name": {"$exists": True, "$nin": [None, ""]},
            }},
            {"$group": {
                "_id": {"$toLower": {"$trim": {"input": "$display_name"}}},
                "count":         {"$sum": 1},
                "record_ids":    {"$push": "$_id"},
                "display_names": {"$addToSet": "$display_name"},
                "folder_paths":  {"$addToSet": "$folder_path"},
            }},
            {"$match": {"count": {"$gt": 1}}},
            {"$sort":  {"count": -1}},
            {"$limit": 5_000},
        ]
        for row in col.aggregate(pipe_b, allowDiskUse=True):
            ids_key = frozenset(str(i) for i in row["record_ids"])
            if ids_key in seen_id_sets:
                continue
            seen_id_sets.add(ids_key)
            groups.append({
                "key":           row["_id"],
                "method":        "name",
                "count":         row["count"],
                "record_ids":    [str(i) for i in row["record_ids"]],
                "display_names": [n for n in (row.get("display_names") or []) if n],
                "folder_paths":  [f for f in (row.get("folder_paths")  or []) if f],
                "duplicates":    row["count"] - 1,
            })

    total        = len(groups)
    extra_total  = sum(g["duplicates"] for g in groups)
    email_groups = sum(1 for g in groups if g["method"] == "email")
    name_groups  = total - email_groups

    start     = (page - 1) * per_page
    page_data = groups[start: start + per_page]

    return jsonify({
        "groups":       page_data,
        "total":        total,
        "extra_total":  extra_total,
        "email_groups": email_groups,
        "name_groups":  name_groups,
        "page":         page,
        "per_page":     per_page,
        "pages":        max(1, -(-total // per_page)),
    })


@app.route("/contacts/merge", methods=["POST"])
def contacts_merge():
    """
    Merge a group of contact records into one.

    Accepts {ids: ["id1", "id2", ...]}.
    The record with the most populated fields is kept; all scalar fields are
    filled in from less-complete duplicates; list/dict fields (email_addresses,
    phone_numbers) are unioned; notes are concatenated.
    The duplicate records are then deleted.
    """
    from bson import ObjectId

    payload = request.get_json(force=True) or {}
    ids     = payload.get("ids") or []

    if len(ids) < 2:
        return jsonify({"error": "Need at least 2 IDs to merge"}), 400

    col = get_col()

    # Resolve IDs (may be ObjectId strings or plain string UUIDs)
    object_ids = []
    for raw in ids:
        try:
            object_ids.append(ObjectId(str(raw)))
        except Exception:
            object_ids.append(raw)

    docs = list(col.find({"_id": {"$in": object_ids}}))
    if len(docs) < 2:
        return jsonify({"error": f"Found only {len(docs)} of {len(ids)} contact records"}), 404

    # Sort descending by field completeness so the richest doc is primary
    def _completeness(doc):
        return sum(1 for v in doc.values() if v not in (None, "", [], {}))

    docs.sort(key=_completeness, reverse=True)
    primary = docs[0]

    # ── Merge scalar fields ───────────────────────────────────────────────────
    SCALAR_FIELDS = [
        "display_name", "first_name", "last_name", "middle_name",
        "honorific", "suffix", "nickname",
        "company_name", "department", "job_title", "office_location",
        "manager", "assistant", "spouse", "children",
        "birthday", "anniversary", "web_page", "im_address", "categories",
        "business_street", "business_street2",
        "business_city",  "business_state",  "business_postal",  "business_country",
        "home_street",    "home_city",    "home_state",    "home_postal",    "home_country",
        "other_street",   "other_city",   "other_state",   "other_postal",   "other_country",
    ]

    merged = dict(primary)

    for doc in docs[1:]:
        # Fill in any empty scalar fields from the secondary doc
        for field in SCALAR_FIELDS:
            if not merged.get(field) and doc.get(field):
                merged[field] = doc[field]

        # Union email_addresses
        seen_emails: set = {e.lower() for e in (merged.get("email_addresses") or [])}
        for e in (doc.get("email_addresses") or []):
            if e and e.lower() not in seen_emails:
                merged.setdefault("email_addresses", []).append(e)
                seen_emails.add(e.lower())
        # Also absorb from_addr of secondary
        fa = (doc.get("from_addr") or "").strip()
        if fa and fa.lower() not in seen_emails:
            merged.setdefault("email_addresses", []).append(fa)
            seen_emails.add(fa.lower())

        # Union phone_numbers dict
        existing_phones: dict = merged.get("phone_numbers") or {}
        for label, number in (doc.get("phone_numbers") or {}).items():
            if label not in existing_phones and number:
                existing_phones[label] = number
        merged["phone_numbers"] = existing_phones

        # Concatenate unique notes
        existing_notes = (merged.get("notes") or merged.get("body_plain") or "").strip()
        new_notes      = (doc.get("notes")    or doc.get("body_plain")    or "").strip()
        if new_notes and new_notes not in existing_notes:
            merged["notes"] = (existing_notes + "\n\n---\n\n" + new_notes).strip() \
                              if existing_notes else new_notes

    # Ensure from_addr reflects primary email
    if not merged.get("from_addr"):
        all_emails = merged.get("email_addresses") or []
        if all_emails:
            merged["from_addr"] = all_emails[0]

    # Rebuild subject/display_name consistency
    merged["subject"] = merged.get("display_name") or merged.get("subject") or ""

    # Track provenance
    merged["merged_from"] = [str(d["_id"]) for d in docs[1:]]
    merged["merged_at"]   = datetime.datetime.utcnow()

    update_fields = {k: v for k, v in merged.items() if k != "_id"}
    col.update_one({"_id": primary["_id"]}, {"$set": update_fields})

    # Delete the duplicates
    dup_ids    = [d["_id"] for d in docs[1:]]
    del_result = col.delete_many({"_id": {"$in": dup_ids}})

    return jsonify({
        "merged":    1,
        "deleted":   del_result.deleted_count,
        "kept_id":   str(primary["_id"]),
        "kept_name": primary.get("display_name") or primary.get("subject") or "",
    })


# Keep /emails as an alias for backwards compatibility
@app.route("/emails")
def all_emails():
    return records()


def _search_text_file(fpath, q_lower, q, snippet_len=200):
    """Return a snippet string if q_lower is found in fpath, else None."""
    try:
        with open(fpath, "r", encoding="utf-8", errors="replace") as fh:
            text = fh.read(500_000)
        idx = text.lower().find(q_lower)
        if idx == -1:
            return None
        start = max(0, idx - snippet_len // 2)
        end   = min(len(text), idx + len(q) + snippet_len // 2)
        return ("…" if start > 0 else "") + \
               " ".join(text[start:end].split()) + \
               ("…" if end < len(text) else "")
    except Exception:
        return None


@app.route("/search-files")
def search_files():
    """
    Search the text content of files inside the Attachments folder tree.
    Plain-text files (txt/html/csv/…) are read directly.
    PDFs are searched via the pdf_text cache (built by /rebuild-pdf-cache).
    Params: q (required), folder (optional subfolder name or 'all')
    Returns up to 200 matches with filename, folder, and a snippet.
    """
    q      = request.args.get("q", "").strip()
    folder = request.args.get("folder", "all").strip()
    if not q:
        return jsonify({"error": "q is required"}), 400

    TEXT_EXTS = {".txt", ".html", ".htm", ".xml", ".json",
                 ".log", ".csv", ".tsv", ".md", ".ini"}
    q_lower  = q.lower()
    results  = []
    searched = 0
    skipped  = 0

    # ── 1. Search plain-text files ──────────────────────────────────────────
    _ad = get_attach_dir()
    roots = ([_ad] if folder == "all"
             else [os.path.join(_ad, folder)])

    for root in roots:
        if not os.path.isdir(root):
            continue
        for dirpath, _, filenames in os.walk(root):
            # Skip internal cache folders
            rel_dir = os.path.relpath(dirpath, _ad).replace(os.sep, "/")
            if rel_dir.split("/")[0] in _INTERNAL_FOLDERS:
                continue
            for fname in filenames:
                ext = os.path.splitext(fname)[1].lower()
                if ext not in TEXT_EXTS:
                    skipped += 1
                    continue
                searched += 1
                snippet = _search_text_file(
                    os.path.join(dirpath, fname), q_lower, q)
                if snippet is not None:
                    results.append({"folder": rel_dir,
                                    "filename": fname, "snippet": snippet})
                    if len(results) >= 200:
                        break
            if len(results) >= 200:
                break

    # ── 2. Search binary-file text caches ───────────────────────────────────
    # Each cache maps:  cache_dir/name.txt  →  real_folder/name.ext
    CACHES = [
        # (cache_dir_fn,        real_folder,  folder_filter,  orig_ext)
        (get_pdf_text_dir,   "pdf",         "pdf",         ".pdf"),
        (get_word_text_dir,  "Word",        "Word",        ".docx"),
        (get_excel_text_dir, "Excel",       "Excel",       ".xlsx"),
        (get_pptx_text_dir,  "PowerPoint",  "PowerPoint",  ".pptx"),
    ]

    cache_counts: dict = {}
    for cache_dir_fn, real_folder, filt, orig_ext in CACHES:
        cache_dir = cache_dir_fn()
        if not os.path.isdir(cache_dir):
            cache_counts[filt] = 0
            continue
        cache_files = [f for f in os.listdir(cache_dir) if f.lower().endswith(".txt")]
        cache_counts[filt] = len(cache_files)
        if len(results) >= 200:
            continue
        if folder not in ("all", filt):
            continue
        for cache_fname in sorted(cache_files):
            searched += 1
            snippet = _search_text_file(
                os.path.join(cache_dir, cache_fname), q_lower, q)
            if snippet is not None:
                # Restore original filename: strip .txt, put back real extension
                base       = os.path.splitext(cache_fname)[0]   # e.g. "invoice"
                orig_fname = base + orig_ext                     # "invoice.pdf"
                # Check the real file exists (might have different ext for Word .doc)
                real_path = os.path.join(_ad, real_folder, orig_fname)
                if not os.path.isfile(real_path):
                    # Fallback: use whatever file has this base name
                    for f in os.listdir(os.path.join(_ad, real_folder)):
                        if os.path.splitext(f)[0] == base:
                            orig_fname = f; break
                results.append({"folder": real_folder,
                                 "filename": orig_fname, "snippet": snippet})
                if len(results) >= 200:
                    break

    pdf_cache_ready = cache_counts.get("pdf", 0) > 0

    # ── 3. Filename matches (files not already in content results) ───────────
    content_keys = {(r["folder"], r["filename"]) for r in results}
    for fn in ["pdf", "Word", "Excel", "PowerPoint", "Images", "Videos", "Text", "Other"]:
        if folder not in ("all", fn):
            continue
        fp = os.path.join(_ad, fn)
        if not os.path.isdir(fp):
            continue
        for fname in os.listdir(fp):
            if q_lower in fname.lower() and os.path.isfile(os.path.join(fp, fname)):
                if (fn, fname) not in content_keys:
                    results.append({"folder": fn, "filename": fname, "snippet": None})

    return jsonify({
        "q":               q,
        "folder":          folder,
        "searched":        searched,
        "skipped":         skipped,
        "total":           len(results),
        "results":         results,
        "pdf_cache_ready": pdf_cache_ready,
        "pdf_cache_count": cache_counts.get("pdf", 0),
        "cache_counts":    cache_counts,
    })


def _extract_file_text(fpath: str) -> str:
    """Extract plain text from a binary attachment file (pdf/docx/xlsx/pptx)."""
    ext = os.path.splitext(fpath)[1].lower()
    try:
        if ext == ".pdf":
            from pdfminer.high_level import extract_text as _pdf_extract
            return (_pdf_extract(fpath) or "").strip()
        if ext == ".docx":
            from docx import Document as _DocxDoc
            doc = _DocxDoc(fpath)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        if ext == ".doc":
            # python-docx can't open legacy .doc — return empty so it's skipped
            return ""
        if ext in (".xlsx", ".xlsm"):
            import openpyxl as _xl
            wb = _xl.load_workbook(fpath, read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    line = " ".join(str(c) for c in row if c is not None)
                    if line.strip():
                        rows.append(line)
            wb.close()
            return "\n".join(rows)
        if ext in (".pptx", ".ppt"):
            from pptx import Presentation as _Prs
            prs = _Prs(fpath)
            parts = []
            for slide in prs.slides:
                for shape in slide.shapes:
                    if hasattr(shape, "text") and shape.text.strip():
                        parts.append(shape.text)
            return "\n".join(parts)
    except Exception:
        pass
    return ""


def _build_cache_for_folder(src_dir: str, cache_dir: str,
                             exts: set, emit, base_pct: int, end_pct: int):
    """Extract text from every matching file in src_dir → cache_dir/*.txt."""
    if not os.path.isdir(src_dir):
        return 0
    files = sorted(f for f in os.listdir(src_dir)
                   if os.path.splitext(f)[1].lower() in exts)
    if not files:
        return 0
    os.makedirs(cache_dir, exist_ok=True)
    for i, fname in enumerate(files, 1):
        pct = base_pct + int((i - 1) / len(files) * (end_pct - base_pct))
        emit(f"[{i}/{len(files)}] {fname[:55]}", pct)
        text = _extract_file_text(os.path.join(src_dir, fname))
        cache_name = os.path.splitext(fname)[0] + ".txt"
        try:
            with open(os.path.join(cache_dir, cache_name),
                      "w", encoding="utf-8", errors="replace") as fh:
                fh.write(text)
        except Exception:
            pass
    return len(files)


@app.route("/rebuild-file-cache")
@app.route("/rebuild-pdf-cache")          # keep old route working
def rebuild_file_cache():
    """SSE stream: build text search cache for PDF, Word, Excel, PowerPoint files."""
    # Capture per-user dirs before thread (no request context inside thread)
    _pdf_dir  = get_pdf_dir();       _pdf_txt  = get_pdf_text_dir()
    _word_dir = _user_attach_subdir("Word");  _word_txt = get_word_text_dir()
    _xl_dir   = _user_attach_subdir("Excel"); _xl_txt   = get_excel_text_dir()
    _pp_dir   = _user_attach_subdir("PowerPoint"); _pp_txt = get_pptx_text_dir()

    def _worker(q: queue.Queue):
        def emit(msg, pct, ok=True):
            q.put(json.dumps({"msg": msg, "pct": pct, "ok": ok}))

        totals = {}
        emit("Starting — indexing all document types…", 1)

        totals["PDF"] = _build_cache_for_folder(
            _pdf_dir, _pdf_txt, {".pdf"}, emit, 2, 50)
        emit(f"PDFs done ({totals['PDF']}). Indexing Word docs…", 50)

        totals["Word"] = _build_cache_for_folder(
            _word_dir, _word_txt, {".docx", ".doc"}, emit, 51, 75)
        emit(f"Word done ({totals['Word']}). Indexing Excel…", 75)

        totals["Excel"] = _build_cache_for_folder(
            _xl_dir, _xl_txt, {".xlsx", ".xlsm", ".xls"}, emit, 76, 90)
        emit(f"Excel done ({totals['Excel']}). Indexing PowerPoint…", 90)

        totals["PowerPoint"] = _build_cache_for_folder(
            _pp_dir, _pp_txt, {".pptx", ".ppt"}, emit, 91, 99)

        summary = "  ".join(f"{k}: {v}" for k, v in totals.items() if v)
        emit(f"Done — {summary}", 100)
        q.put(None)

    q: queue.Queue = queue.Queue()
    threading.Thread(target=_worker, args=(q,), daemon=True).start()

    def generate():
        while True:
            try:
                item = q.get(timeout=120)
            except queue.Empty:
                yield "data: \n\n"; continue
            if item is None: break
            yield f"data: {item}\n\n"

    return Response(stream_with_context(generate()),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


def _effective_attach_dir() -> str:
    """Return the per-user attachment directory."""
    return get_attach_dir()


@app.route("/attach-folders")
def attach_folders():
    """List subfolders and file counts inside the user's Attachments directory."""
    _ad = _effective_attach_dir()
    folders = []
    if os.path.isdir(_ad):
        for name in sorted(os.listdir(_ad)):
            fpath = os.path.join(_ad, name)
            if os.path.isdir(fpath) and name not in _INTERNAL_FOLDERS:
                count = sum(len(files) for _, _, files in os.walk(fpath))
                folders.append({"name": name, "count": count})
    return jsonify(folders)


@app.route("/list-files")
def list_files():
    """
    Return metadata for every file inside the Attachments folder.
    Query params:
      folder  – subfolder name or 'all' (default)
      sort    – 'date' | 'name' | 'size' | 'folder'  (default: date)
      order   – 'asc'  | 'desc'                       (default: desc)
    Returns up to 2 000 entries.
    """
    folder  = request.args.get("folder", "all").strip()
    sort_by = request.args.get("sort",   "date").strip()
    order   = request.args.get("order",  "desc").strip()

    _ad = _effective_attach_dir()
    roots = ([_ad] if folder == "all"
             else [os.path.join(_ad, folder)])

    files = []
    all_fnames = []
    all_fpaths = []
    for root in roots:
        if not os.path.isdir(root):
            continue
        for dirpath, _, filenames in os.walk(root):
            rel_dir = os.path.relpath(dirpath, _ad).replace(os.sep, "/")
            if rel_dir.split("/")[0] in _INTERNAL_FOLDERS:
                continue
            for fname in filenames:
                fpath = os.path.join(dirpath, fname)
                try:
                    st   = os.stat(fpath)
                    sz   = st.st_size
                    if   sz < 1_024:         size_str = f"{sz} B"
                    elif sz < 1_048_576:     size_str = f"{sz/1_024:.1f} KB"
                    elif sz < 1_073_741_824: size_str = f"{sz/1_048_576:.1f} MB"
                    else:                    size_str = f"{sz/1_073_741_824:.2f} GB"
                    files.append({
                        "folder":      rel_dir,
                        "filename":    fname,
                        "_fpath":      fpath,   # stripped before response
                        "size_bytes":  sz,
                        "size_str":    size_str,
                        "email_ts":    None,    # filled from MongoDB below
                        "email_str":   "",
                    })
                    all_fnames.append(fname)
                    all_fpaths.append(fpath)
                except Exception:
                    pass

    # Build date maps from MongoDB:
    #   1. disk_path → date  (exact match; most reliable when paths haven't changed)
    #   2. filename  → date  (fallback for cross-platform imports where paths differ)
    # We try the app's configured collection first.  If it yields nothing we also
    # try the legacy local database (mydb.pst_items on localhost) which holds data
    # imported before the multi-user / remote-MongoDB migration.
    path_to_date:  dict = {}
    fname_to_date: dict = {}

    _fpath_set  = set(all_fpaths)
    _fname_set  = set(all_fnames)

    def _fill_date_maps(col):
        """Populate path_to_date and fname_to_date from a collection."""
        if not _fpath_set and not _fname_set:
            return
        try:
            # Push the filter into MongoDB so it only scans matching docs
            match_stage = {"$or": []}
            if _fpath_set:
                match_stage["$or"].append(
                    {"attachments.disk_path": {"$in": list(_fpath_set)}})
            if _fname_set:
                match_stage["$or"].append(
                    {"attachments.filename": {"$in": list(_fname_set)}})
            for row in col.aggregate([
                {"$match":   match_stage},
                {"$unwind":  "$attachments"},
                {"$project": {"_id": 0,
                               "dp": "$attachments.disk_path",
                               "fn": "$attachments.filename",
                               "d":  "$date"}},
            ], allowDiskUse=True):
                d  = row.get("d")
                if not d:
                    continue
                dp = row.get("dp")
                fn = row.get("fn")
                if dp and dp in _fpath_set:
                    if dp not in path_to_date or d < path_to_date[dp]:
                        path_to_date[dp] = d
                if fn and fn in _fname_set:
                    if fn not in fname_to_date or d < fname_to_date[fn]:
                        fname_to_date[fn] = d
        except Exception:
            pass

    if all_fnames or all_fpaths:
        # Primary: configured MongoDB collection
        _fill_date_maps(get_col())

        # Fallback: if we still have files with no date, try the legacy local DB
        unmatched_paths  = {f["_fpath"]   for f in files if f["_fpath"] not in path_to_date}
        unmatched_fnames = {f["filename"] for f in files if f["filename"] not in fname_to_date}
        if unmatched_paths or unmatched_fnames:
            try:
                _legacy_col = MongoClient(
                    "mongodb://localhost:27017",
                    serverSelectionTimeoutMS=1500
                )["mydb"]["pst_items"]
                _fill_date_maps(_legacy_col)
            except Exception:
                pass

    for f in files:
        fp = f.pop("_fpath")
        dt = path_to_date.get(fp) or fname_to_date.get(f["filename"])
        if dt:
            f["email_ts"]  = dt.timestamp()
            f["email_str"] = dt.strftime("%Y-%m-%d %H:%M")
        else:
            try:
                mtime = os.stat(fp).st_mtime
            except Exception:
                mtime = 0.0
            f["email_ts"]  = mtime
            f["email_str"] = datetime.datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")

    # Server-side sort
    key_map = {
        "date":   lambda f: f["email_ts"] or 0,
        "name":   lambda f: f["filename"].lower(),
        "size":   lambda f: f["size_bytes"],
        "folder": lambda f: f["folder"].lower(),
    }
    files.sort(key=key_map.get(sort_by, key_map["date"]),
               reverse=(order == "desc"))

    total = len(files)
    MAX   = 2_000
    return jsonify({
        "folder":   folder,
        "sort":     sort_by,
        "order":    order,
        "total":    total,
        "returned": min(total, MAX),
        "files":    files[:MAX],
    })


@app.route("/local-file/<path:rel_path>")
def local_file(rel_path):
    """
    Serve a file from the Attachments directory by its relative path.
    Viewable types open inline; add ?dl=1 to force download.
    """
    _ad = get_attach_dir()
    safe_base = os.path.realpath(_ad)
    target    = os.path.realpath(os.path.join(_ad, rel_path))
    # Guard against path-traversal
    if not (target == safe_base or target.startswith(safe_base + os.sep)):
        return jsonify({"error": "Invalid path"}), 400
    if not os.path.isfile(target):
        return jsonify({"error": "File not found"}), 404

    directory = os.path.dirname(target)
    filename  = os.path.basename(target)
    force_dl  = request.args.get("dl") == "1"

    ct, _ = mimetypes.guess_type(filename)
    ct    = ct or "application/octet-stream"

    INLINE = {
        "application/pdf", "application/json",
        "text/plain", "text/html", "text/csv", "text/xml",
    }
    inline = (not force_dl) and (
        ct in INLINE or ct.startswith("image/")
        or ct.startswith("text/") or ct.startswith("video/")
    )
    return send_from_directory(directory, filename,
                               as_attachment=not inline,
                               download_name=filename)


@app.route("/clear-files", methods=["POST"])
def clear_files():
    """Delete all files in the selected Attachments subfolders (keeps the folders themselves)."""
    data    = request.get_json(silent=True) or {}
    folders = data.get("folders", [])
    if not folders:
        return jsonify({"error": "No folders specified"}), 400

    _ad     = get_attach_dir()
    deleted = 0
    errors  = 0
    for folder in folders:
        # Safety: only allow direct children of user's attach dir
        safe_base = os.path.realpath(_ad)
        target    = os.path.realpath(os.path.join(_ad, folder))
        if not target.startswith(safe_base + os.sep) or not os.path.isdir(target):
            errors += 1
            continue
        for dirpath, _, filenames in os.walk(target):
            for fname in filenames:
                try:
                    os.remove(os.path.join(dirpath, fname))
                    deleted += 1
                except Exception:
                    errors += 1

    return jsonify({"deleted": deleted, "errors": errors})


@app.route("/export-info")
def export_info():
    """Return file count and total uncompressed size for selected folders."""
    _ad     = get_attach_dir()
    folders = request.args.getlist("folder")
    roots   = ([_ad] if not folders or "all" in folders
               else [os.path.join(_ad, f) for f in folders])
    total_files = 0
    total_bytes = 0
    for root in roots:
        if not os.path.isdir(root):
            continue
        for dirpath, _, filenames in os.walk(root):
            for fname in filenames:
                try:
                    total_bytes += os.path.getsize(os.path.join(dirpath, fname))
                    total_files += 1
                except Exception:
                    pass
    if   total_bytes < 1_048_576:     size_str = f"{total_bytes/1_024:.1f} KB"
    elif total_bytes < 1_073_741_824: size_str = f"{total_bytes/1_048_576:.1f} MB"
    else:                             size_str = f"{total_bytes/1_073_741_824:.2f} GB"
    return jsonify({"files": total_files, "bytes": total_bytes, "size_str": size_str})


@app.route("/export-zip")
def export_zip():
    """Stream a ZIP archive of the selected attachment folders."""
    import zipfile
    _ad     = get_attach_dir()
    folders = request.args.getlist("folder")
    roots   = ([_ad] if not folders or "all" in folders
               else [os.path.join(_ad, f) for f in folders])

    def generate():
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for root in roots:
                if not os.path.isdir(root):
                    continue
                for dirpath, _, filenames in os.walk(root):
                    rel_dir = os.path.relpath(dirpath, _ad).replace(os.sep, "/")
                    for fname in filenames:
                        fpath   = os.path.join(dirpath, fname)
                        arcname = f"{rel_dir}/{fname}" if rel_dir != "." else fname
                        try:
                            zf.write(fpath, arcname)
                        except Exception:
                            pass
        buf.seek(0)
        while True:
            chunk = buf.read(65_536)
            if not chunk:
                break
            yield chunk

    return Response(
        stream_with_context(generate()),
        mimetype="application/zip",
        headers={"Content-Disposition": 'attachment; filename="attachments_export.zip"'},
    )


# ---------------------------------------------------------------------------
# CSV Contacts Import
# ---------------------------------------------------------------------------

# Normalised (lower-stripped) Outlook CSV column → internal field name.
# Handles both classic Outlook (E-mail Address) and new Outlook (Email Address).
_CSV_COL_MAP = {
    # Display / full name
    "name":                      "display_name",
    "full name":                 "display_name",
    "display name":              "display_name",
    # Name parts
    "first name":                "first_name",
    "last name":                 "last_name",
    "middle name":               "middle_name",
    "title":                     "honorific",
    "suffix":                    "suffix",
    "nickname":                  "nickname",
    # E-mail (both hyphenated and space variants)
    "e-mail address":            "email_1",
    "email address":             "email_1",
    "e-mail 2 address":          "email_2",
    "email 2 address":           "email_2",
    "e-mail 3 address":          "email_3",
    "email 3 address":           "email_3",
    # Phone
    "business phone":            "phone_business",
    "business phone 2":          "phone_business2",
    "home phone":                "phone_home",
    "home phone 2":              "phone_home2",
    "mobile phone":              "phone_mobile",
    "car phone":                 "phone_car",
    "other phone":               "phone_other",
    "primary phone":             "phone_primary",
    "pager":                     "phone_pager",
    "business fax":              "phone_fax_business",
    "home fax":                  "phone_fax_home",
    # Company
    "company":                   "company_name",
    "department":                "department",
    "job title":                 "job_title",
    "office location":           "office_location",
    "manager's name":            "manager",
    "assistant's name":          "assistant",
    # Business address
    "business street":           "business_street",
    "business street 2":         "business_street2",
    "business city":             "business_city",
    "business state":            "business_state",
    "business postal code":      "business_postal",
    "business country/region":   "business_country",
    # Home address
    "home street":               "home_street",
    "home city":                 "home_city",
    "home state":                "home_state",
    "home postal code":          "home_postal",
    "home country/region":       "home_country",
    # Other address
    "other street":              "other_street",
    "other city":                "other_city",
    "other state":               "other_state",
    "other postal code":         "other_postal",
    "other country/region":      "other_country",
    # Personal / misc
    "birthday":                  "birthday",
    "anniversary":               "anniversary",
    "spouse":                    "spouse",
    "children":                  "children",
    "notes":                     "notes",
    "web page":                  "web_page",
    "im address":                "im_address",
    "categories":                "categories",
    "account":                   "account",
    "referred by":               "referred_by",
}


def _csv_row_to_doc(row: dict, header_map: dict):
    """Convert one CSV row (already filtered through header_map) to a MongoDB doc."""
    mapped = {}
    for csv_col, field in header_map.items():
        val = (row.get(csv_col) or "").strip()
        if val:
            mapped[field] = val

    # Build display name from parts when no direct "Name" column
    display = mapped.get("display_name") or " ".join(filter(None, [
        mapped.get("first_name", ""),
        mapped.get("middle_name", ""),
        mapped.get("last_name", ""),
    ])).strip() or mapped.get("email_1", "")

    if not display:
        return None   # skip empty rows

    emails = [mapped.get(f"email_{i}") for i in range(1, 4)]
    emails = [e for e in emails if e]

    phone_map = {}
    for label, field in [
        ("mobile",   "phone_mobile"),
        ("business", "phone_business"),
        ("home",     "phone_home"),
        ("fax",      "phone_fax_business"),
        ("other",    "phone_other"),
    ]:
        if field in mapped:
            phone_map[label] = mapped[field]

    doc = {
        "_id":             str(uuid.uuid4()),
        "item_type":       "contact",
        "subject":         display,
        "display_name":    display,
        "folder_path":     "Contacts/CSV Import",
        "from_addr":       emails[0] if emails else "",
        "to_addrs":        [],
        "cc_addrs":        [],
        "body_plain":      mapped.get("notes", ""),
        "has_attachments": False,
        "attachments":     [],
        "imported_at":     datetime.datetime.utcnow(),
        "source":          "csv",
    }

    # Carry every other mapped field into the document
    for key in (
        "first_name", "last_name", "middle_name", "honorific", "suffix", "nickname",
        "company_name", "department", "job_title", "office_location",
        "manager", "assistant", "spouse", "children",
        "birthday", "anniversary", "web_page", "im_address", "categories",
        "account", "referred_by", "notes",
        "business_street", "business_street2", "business_city",
        "business_state",  "business_postal",  "business_country",
        "home_street", "home_city", "home_state", "home_postal", "home_country",
        "other_street", "other_city", "other_state", "other_postal", "other_country",
    ):
        if key in mapped:
            doc[key] = mapped[key]

    if emails:
        doc["email_addresses"] = emails
    if phone_map:
        doc["phone_numbers"] = phone_map

    return doc


def _run_csv_import(job_id: str, csv_path: str, q: queue.Queue):
    """Parse an Outlook contacts CSV and insert docs into MongoDB."""
    import csv

    try:
        # Try common encodings; Outlook classic uses cp1252/latin-1 with BOM
        encoding = "utf-8-sig"
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                with open(csv_path, "r", encoding=enc, newline="") as fh:
                    fh.read(512)
                encoding = enc
                break
            except (UnicodeDecodeError, LookupError):
                continue

        with open(csv_path, "r", encoding=encoding, newline="") as fh:
            reader  = csv.DictReader(fh)
            headers = reader.fieldnames or []
            rows    = list(reader)

        # Build header_map: original_csv_header → internal_field
        header_map = {}
        for h in headers:
            norm = h.strip().lower()
            if norm in _CSV_COL_MAP:
                header_map[h] = _CSV_COL_MAP[norm]

        total = len(rows)
        q.put(f"Found {total:,} row(s) — encoding={encoding}, "
              f"{len(header_map)}/{len(headers)} columns recognised")

        if total == 0:
            q.put("No contacts found in file.")
            jobs[job_id]["status"] = "done"
            q.put(None)
            return

        col      = get_col()
        inserted = 0
        skipped  = 0
        batch    = []
        BATCH    = 100

        for i, row in enumerate(rows, 1):
            doc = _csv_row_to_doc(row, header_map)
            if doc is None:
                skipped += 1
                continue
            batch.append(doc)
            if len(batch) >= BATCH:
                col.insert_many(batch, ordered=False)
                inserted += len(batch)
                batch = []
                q.put(f"Flushed {inserted:,} contacts…")

        if batch:
            col.insert_many(batch, ordered=False)
            inserted += len(batch)

        q.put(f"Flushed {inserted:,} contacts")
        q.put(f"Done — {inserted:,} contacts imported, {skipped} row(s) skipped.")
        jobs[job_id]["status"] = "done"

    except Exception as e:
        q.put(f"ERROR: {e}")
        jobs[job_id]["status"] = "error"
    finally:
        q.put(None)   # sentinel


@app.route("/upload-csv", methods=["POST"])
def upload_csv():
    """Accept an Outlook contacts CSV, kick off a background import job."""
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400
    if not f.filename.lower().endswith(".csv"):
        return jsonify({"error": "Only .csv files are supported"}), 400

    save_path = os.path.join(CSV_UPLOAD_DIR, f.filename)
    f.save(save_path)

    job_id = str(uuid.uuid4())
    q      = queue.Queue()
    jobs[job_id] = {"queue": q, "status": "running", "filename": f.filename}

    threading.Thread(target=_run_csv_import, args=(job_id, save_path, q),
                     daemon=True).start()
    return jsonify({"job_id": job_id, "filename": f.filename})


@app.route("/preview-text/<path:rel_path>")
def preview_text(rel_path):
    """Return cached or live-extracted text for an office-format attachment."""
    _ad       = get_attach_dir()
    safe_base = os.path.realpath(_ad)
    target    = os.path.realpath(os.path.join(_ad, rel_path))
    if not (target == safe_base or target.startswith(safe_base + os.sep)):
        return jsonify({"error": "Invalid path"}), 400
    if not os.path.isfile(target):
        return jsonify({"error": "File not found"}), 404

    filename  = os.path.basename(target)
    base_name = os.path.splitext(filename)[0]
    ext       = os.path.splitext(filename)[1].lower()

    cache_dirs = {
        ".pdf":  get_pdf_text_dir(),
        ".docx": get_word_text_dir(), ".doc": get_word_text_dir(),
        ".xlsx": get_excel_text_dir(), ".xlsm": get_excel_text_dir(), ".xls": get_excel_text_dir(),
        ".pptx": get_pptx_text_dir(),  ".ppt": get_pptx_text_dir(),
    }

    # Try cache first (fast)
    cache_dir = cache_dirs.get(ext)
    if cache_dir:
        cache_path = os.path.join(cache_dir, base_name + ".txt")
        if os.path.isfile(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8", errors="replace") as fh:
                    return jsonify({"text": fh.read(200_000), "source": "cache"})
            except Exception:
                pass

    # Fall back to live extraction
    text = _extract_file_text(target)
    if text:
        return jsonify({"text": text[:200_000], "source": "live"})
    return jsonify({"text": None})


@app.route("/stats")
def stats():
    col   = get_col()
    total = col.count_documents({})
    return jsonify({"total": total, "db": col.database.name, "collection": COLLECTION})


@app.route("/attachment/<file_id>")
def download_attachment(file_id):
    """Stream an attachment from GridFS.
    Viewable types (PDF, images, plain text, HTML) are served inline so the
    browser can open them directly.  Everything else (Office docs, zip, …)
    is sent as a forced download.
    """
    INLINE_TYPES = {
        "application/pdf",
        "text/plain", "text/html", "text/csv", "text/xml",
        "image/jpeg", "image/jpg", "image/png", "image/gif",
        "image/svg+xml", "image/webp", "image/bmp", "image/tiff",
    }
    try:
        fs       = get_fs()
        grid_out = fs.get(ObjectId(file_id))
        data     = grid_out.read()
        filename = grid_out.filename or "attachment"
        ct       = grid_out.content_type or "application/octet-stream"
        # ?dl=1 forces a file download regardless of type
        force_dl = request.args.get("dl") == "1"
        inline   = (not force_dl) and (
            ct in INLINE_TYPES or ct.startswith("image/") or ct.startswith("text/")
        )
        return send_file(
            io.BytesIO(data),
            mimetype=ct,
            as_attachment=not inline,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 404


@app.route("/save-attachment/<file_id>", methods=["POST"])
def save_attachment(file_id):
    """Save a GridFS attachment to the local Attachments folder tree."""
    try:
        fs       = get_fs()
        grid_out = fs.get(ObjectId(file_id))
        data     = grid_out.read()
        filename = grid_out.filename or "attachment"

        folder   = _attach_folder(filename)

        # Avoid overwriting: append a counter if the file already exists
        dest = os.path.join(folder, filename)
        if os.path.exists(dest):
            base, ext = os.path.splitext(filename)
            n = 1
            while os.path.exists(dest):
                dest = os.path.join(folder, f"{base}_{n}{ext}")
                n += 1

        with open(dest, "wb") as fh:
            fh.write(data)

        return jsonify({
            "ok":       True,
            "saved_to": dest,
            "filename": os.path.basename(dest),
            "folder":   os.path.basename(folder),
            "bytes":    len(data),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/clear", methods=["POST"])
def clear_collection():
    """Drop the entire collection + GridFS + Attachments folder, ready for a fresh import."""
    db     = get_db()
    col    = db[COLLECTION]
    before = col.count_documents({})
    col.drop()

    # Drop GridFS collections
    db["fs.files"].drop()
    db["fs.chunks"].drop()

    # Recreate standard indexes on the fresh collection
    col.create_index("folder_path")
    col.create_index("from_addr")
    col.create_index("date")
    col.create_index("subject")
    col.create_index("message_id")
    col.create_index("item_type")

    # Wipe Attachments folder and recreate it empty
    import shutil
    attach_dir = get_attach_dir()
    if os.path.isdir(attach_dir):
        shutil.rmtree(attach_dir)
    os.makedirs(attach_dir, exist_ok=True)

    # Delete any stale .part upload files so they don't fill the disk
    upload_dir = get_upload_dir()
    for f in os.listdir(upload_dir):
        if f.endswith(".part"):
            os.remove(os.path.join(upload_dir, f))

    return jsonify({"ok": True, "deleted": before})


# ---------------------------------------------------------------------------
# Data Export
# ---------------------------------------------------------------------------
import csv as _csv_mod


def _csv_stream(headers, row_generator):
    """Yield CSV text line-by-line (suitable for a streamed Flask Response)."""
    buf = io.StringIO()
    writer = _csv_mod.writer(buf)
    writer.writerow(headers)
    yield buf.getvalue(); buf.seek(0); buf.truncate()
    for row in row_generator:
        writer.writerow(row)
        yield buf.getvalue(); buf.seek(0); buf.truncate()


def _build_record_query(args):
    """Build a MongoDB query dict from common export filter args."""
    item_type = args.get("type",       "all").strip().lower()
    q         = args.get("q",          "").strip()
    date_from = args.get("date_from",  "").strip()
    date_to   = args.get("date_to",    "").strip()
    from_addr = args.get("from_addr",  "").strip()
    to_addr   = args.get("to_addr",    "").strip()
    folder_q  = args.get("folder_path","").strip()

    query: dict = {}
    if item_type and item_type != "all":
        if item_type == "hasatt":
            query["has_attachments"] = True
        else:
            query["item_type"] = item_type
    if from_addr:
        query["from_addr"] = {"$regex": from_addr, "$options": "i"}
    if to_addr:
        rx = {"$regex": to_addr, "$options": "i"}
        query["$or"] = [{"to_addrs": rx}, {"cc_addrs": rx}]
    if folder_q:
        query["folder_path"] = {"$regex": folder_q, "$options": "i"}

    date_filter: dict = {}
    if date_from:
        try: date_filter["$gte"] = datetime.datetime.strptime(date_from, "%Y-%m-%d")
        except ValueError: pass
    if date_to:
        try:
            date_filter["$lte"] = (
                datetime.datetime.strptime(date_to, "%Y-%m-%d")
                + datetime.timedelta(days=1)
            )
        except ValueError: pass
    if date_filter:
        query["date"] = date_filter

    return query, q   # return q separately so caller can add $text


def _export_filename(args, ext):
    """Build a descriptive export filename from the active filter args."""
    import re as _re
    parts = []
    item_type = args.get("type", "all").strip().lower()
    if item_type and item_type != "all":
        parts.append(item_type if item_type != "hasatt" else "with_attachments")
    else:
        parts.append("records")
    q = args.get("q", "").strip()
    if q:
        slug = _re.sub(r"[^\w]+", "_", q)[:24].strip("_")
        if slug:
            parts.append(slug)
    if args.get("date_from"):
        parts.append("from_" + args["date_from"])
    if args.get("date_to"):
        parts.append("to_" + args["date_to"])
    if args.get("from_addr"):
        slug = _re.sub(r"[^\w@.]+", "_", args["from_addr"])[:20].strip("_")
        parts.append(slug)
    base = "_".join(parts) or "export"
    return f"{base}.{ext}"


@app.route("/export/emails.csv")
def export_emails_csv():
    """Stream all matching records as a UTF-8-BOM CSV download."""
    col   = get_col()
    limit = min(int(request.args.get("limit", 100_000)), 500_000)
    query, q = _build_record_query(request.args)
    if q:
        _ensure_text_index(col)
        query["$text"] = {"$search": q}

    sort_by = request.args.get("sort", "date")
    order   = int(request.args.get("order", -1))
    if sort_by not in {"date", "subject", "from_addr"}:
        sort_by = "date"
    if order not in (-1, 1):
        order = -1

    proj = {
        "date": 1, "from_addr": 1, "to_addrs": 1, "cc_addrs": 1,
        "subject": 1, "folder_path": 1, "item_type": 1,
        "has_attachments": 1, "body_plain": 1,
    }
    cursor = col.find(query, proj).sort(sort_by, order).limit(limit)

    col_headers = ["Date", "From", "To", "CC", "Subject",
                   "Folder", "Type", "Has Attachments", "Body (first 500 chars)"]

    def rows():
        for doc in cursor:
            body = " ".join((doc.get("body_plain") or "").split())[:500]
            yield [
                doc["date"].strftime("%Y-%m-%d %H:%M") if doc.get("date") else "",
                doc.get("from_addr") or "",
                "; ".join(doc.get("to_addrs") or []),
                "; ".join(doc.get("cc_addrs") or []),
                doc.get("subject") or "",
                doc.get("folder_path") or "",
                doc.get("item_type") or "email",
                "Yes" if doc.get("has_attachments") else "No",
                body,
            ]

    fname = _export_filename(request.args, "csv")
    return Response(
        stream_with_context(_csv_stream(col_headers, rows())),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.route("/export/emails.xlsx")
def export_emails_xlsx():
    """Export matching records as a formatted Excel workbook (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed — run: pip install openpyxl"}), 500

    col   = get_col()
    limit = min(int(request.args.get("limit", 100_000)), 200_000)
    query, q = _build_record_query(request.args)
    if q:
        _ensure_text_index(col)
        query["$text"] = {"$search": q}

    sort_by = request.args.get("sort",  "date")
    order   = int(request.args.get("order", -1))
    if sort_by not in {"date", "subject", "from_addr"}:
        sort_by = "date"
    if order not in (-1, 1):
        order = -1

    proj = {
        "date": 1, "from_addr": 1, "to_addrs": 1, "cc_addrs": 1,
        "subject": 1, "folder_path": 1, "item_type": 1,
        "has_attachments": 1, "body_plain": 1,
    }
    cursor = col.find(query, proj).sort(sort_by, order).limit(limit)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Records"

    HEADERS = ["Date", "From", "To", "CC", "Subject",
               "Folder", "Type", "Has Attachments", "Body (first 500 chars)"]

    # Style header row
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1A1F36")
    hdr_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18

    for doc in cursor:
        body = " ".join((doc.get("body_plain") or "").split())[:500]
        ws.append([
            doc["date"].strftime("%Y-%m-%d %H:%M") if doc.get("date") else "",
            doc.get("from_addr") or "",
            "; ".join(doc.get("to_addrs") or []),
            "; ".join(doc.get("cc_addrs") or []),
            doc.get("subject") or "",
            doc.get("folder_path") or "",
            doc.get("item_type") or "email",
            "Yes" if doc.get("has_attachments") else "No",
            body,
        ])

    # Auto-filter + column widths
    ws.auto_filter.ref = ws.dimensions
    for ci, w in enumerate([18, 32, 32, 22, 50, 32, 12, 16, 65], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = _export_filename(request.args, "xlsx")
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.route("/process-status")
def process_status():
    """
    Return the status of each PST processing step for the current user.
    Steps: import, virus_scan, indexing, ocr
    Status values: pending | running | done | error
    """
    col = get_col()

    # ── Step 1 & 2: Import + Virus Scan ──────────────────────────────────────
    # Find the most recent job for this user by scanning jobs dict
    # Jobs don't store user, so we fall back to record count heuristic
    user_jobs = [
        j for j in jobs.values()
        if "status" in j and "filename" in j
    ]
    latest_job = user_jobs[-1] if user_jobs else None

    record_count = 0
    try:
        record_count = col.count_documents({})
    except Exception:
        pass

    if record_count == 0 and (not latest_job or latest_job.get("status") in ("done", "error")):
        import_status = "pending"
        virus_status  = "pending"
    elif latest_job:
        job_status = latest_job.get("status", "running")
        import_status   = "done" if job_status == "done" else ("error" if job_status == "error" else "running")
        virus_status    = "done" if job_status in ("done", "error") else "running"
    elif record_count > 0:
        import_status   = "done"
        virus_status    = "done"
    else:
        import_status   = "pending"
        virus_status    = "pending"

    # ── Step 3: Indexing ─────────────────────────────────────────────────────
    index_status = "pending"
    if import_status == "done":
        try:
            idx_info = col.index_information()
            has_text_idx = any(
                any(t == "text" for _, t in v.get("key", []))
                for v in idx_info.values()
            )
            index_status = "done" if has_text_idx else "running"
        except Exception:
            index_status = "pending"

    # ── Step 4: OCR ──────────────────────────────────────────────────────────
    ocr_status_val = "pending"
    if import_status == "done":
        try:
            _pdf_dir = get_pdf_dir()
            _txt_dir = get_pdf_text_dir()
            if os.path.isdir(_pdf_dir):
                pdfs  = [f for f in os.listdir(_pdf_dir) if f.lower().endswith(".pdf")]
                total = len(pdfs)
                if total == 0:
                    ocr_status_val = "done"
                else:
                    done_count = sum(
                        1 for f in pdfs
                        if os.path.isfile(os.path.join(_txt_dir, os.path.splitext(f)[0] + ".txt"))
                    ) if os.path.isdir(_txt_dir) else 0
                    if done_count >= total:
                        ocr_status_val = "done"
                    elif done_count > 0:
                        ocr_status_val = "running"
                    else:
                        ocr_status_val = "pending"
            else:
                ocr_status_val = "pending"
        except Exception:
            ocr_status_val = "pending"

    return jsonify({
        "record_count": record_count,
        "steps": [
            {"id": "virus_scan", "label": "Virus scan",
             "detail": "File scanned and cleared" if virus_status == "done" else "Scanning your PST file for viruses",
             "status": virus_status},
            {"id": "import",     "label": "Import data",
             "detail": f"{record_count:,} records imported" if import_status == "done" else "Importing emails, contacts, and attachments from your PST file",
             "status": import_status},
            {"id": "indexing",   "label": "Indexing",
             "detail": "Full-text index ready" if index_status == "done" else "Building search index across all emails",
             "status": index_status},
            {"id": "ocr",        "label": "Attachment OCR",
             "detail": "Documents are searchable" if ocr_status_val == "done" else "Processing scanned PDFs so you can search within documents",
             "status": ocr_status_val},
        ]
    })


@app.route("/process-flow")
def process_flow():
    return render_template("process_flow.html")


@app.route("/export/emails.eml.zip")
def export_emails_eml_zip():
    """Export all matching records as individual .eml files inside a ZIP archive."""
    import zipfile

    col   = get_col()
    limit = min(int(request.args.get("limit", 10_000)), 10_000)
    query, q = _build_record_query(request.args)
    if q:
        _ensure_text_index(col)
        query["$text"] = {"$search": q}

    sort_by = request.args.get("sort", "date")
    order   = int(request.args.get("order", -1))
    if sort_by not in {"date", "subject", "from_addr"}:
        sort_by = "date"
    if order not in (-1, 1):
        order = -1

    proj = {
        "date": 1, "from_addr": 1, "to_addrs": 1, "cc_addrs": 1,
        "subject": 1, "folder_path": 1, "body_plain": 1, "body_html": 1,
        "message_id": 1,
    }
    docs = list(col.find(query, proj).sort(sort_by, order).limit(limit))

    buf = io.BytesIO()
    seen_names: dict = {}

    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for i, doc in enumerate(docs, 1):
            base = _eml_filename(doc)
            name = base
            if name in seen_names:
                stem, ext = name.rsplit(".", 1)
                name = f"{stem}_{i}.{ext}"
            seen_names[name] = True
            try:
                eml_bytes = _build_eml(doc)
            except Exception:
                eml_bytes = b""
            zf.writestr(name, eml_bytes)

    buf.seek(0)
    fname = _export_filename(request.args, "eml.zip")
    return Response(
        buf.read(),
        mimetype="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.route("/export/contacts.csv")
def export_contacts_csv():
    """Export contacts in Outlook-compatible CSV format (re-importable)."""
    col = get_col()
    query = {"item_type": "contact"}
    q_txt = request.args.get("q", "").strip()
    if q_txt:
        query["$or"] = [
            {"display_name": {"$regex": q_txt, "$options": "i"}},
            {"from_addr":    {"$regex": q_txt, "$options": "i"}},
            {"company_name": {"$regex": q_txt, "$options": "i"}},
        ]
    limit  = min(int(request.args.get("limit", 100_000)), 500_000)
    cursor = col.find(query).sort("display_name", 1).limit(limit)

    # Outlook-compatible column names (matches the import _CSV_COL_MAP in reverse)
    headers = [
        "First Name", "Last Name", "Middle Name", "Title", "Suffix", "Nickname",
        "Name", "E-mail Address", "E-mail 2 Address", "E-mail 3 Address",
        "Business Phone", "Home Phone", "Mobile Phone", "Business Fax",
        "Company", "Department", "Job Title", "Office Location",
        "Business Street", "Business City", "Business State",
        "Business Postal Code", "Business Country/Region",
        "Home Street", "Home City", "Home State", "Home Postal Code", "Home Country/Region",
        "Web Page", "Notes",
    ]

    def rows():
        for doc in cursor:
            emails = doc.get("email_addresses") or [doc.get("from_addr") or ""]
            phones = doc.get("phone_numbers") or {}
            yield [
                doc.get("first_name", ""),
                doc.get("last_name",  ""),
                doc.get("middle_name",""),
                doc.get("honorific",  ""),
                doc.get("suffix",     ""),
                doc.get("nickname",   ""),
                doc.get("display_name",""),
                emails[0] if len(emails) > 0 else "",
                emails[1] if len(emails) > 1 else "",
                emails[2] if len(emails) > 2 else "",
                phones.get("business", ""),
                phones.get("home",     ""),
                phones.get("mobile",   ""),
                phones.get("fax",      ""),
                doc.get("company_name",     ""),
                doc.get("department",       ""),
                doc.get("job_title",        ""),
                doc.get("office_location",  ""),
                doc.get("business_street",  ""),
                doc.get("business_city",    ""),
                doc.get("business_state",   ""),
                doc.get("business_postal",  ""),
                doc.get("business_country", ""),
                doc.get("home_street",  ""),
                doc.get("home_city",    ""),
                doc.get("home_state",   ""),
                doc.get("home_postal",  ""),
                doc.get("home_country", ""),
                doc.get("web_page", ""),
                doc.get("notes",    ""),
            ]

    return Response(
        stream_with_context(_csv_stream(headers, rows())),
        mimetype="text/csv; charset=utf-8-sig",   # BOM so Excel auto-detects UTF-8
        headers={"Content-Disposition": 'attachment; filename="contacts_export.csv"'},
    )


@app.route("/export/contacts.vcf")
def export_contacts_vcf():
    """Export contacts as vCard 3.0 (.vcf) — importable into Outlook, Gmail, Apple Mail."""
    col    = get_col()
    q_txt  = request.args.get("q", "").strip()
    query  = {"item_type": "contact"}
    if q_txt:
        query["$or"] = [
            {"display_name": {"$regex": q_txt, "$options": "i"}},
            {"from_addr":    {"$regex": q_txt, "$options": "i"}},
            {"company_name": {"$regex": q_txt, "$options": "i"}},
        ]
    limit  = min(int(request.args.get("limit", 100_000)), 500_000)
    cursor = col.find(query).sort("display_name", 1).limit(limit)

    def _vf(s: str) -> str:
        """Escape vCard special characters."""
        return (s or "").replace("\\", "\\\\").replace(",", "\\,").replace("\n", "\\n")

    def generate():
        for doc in cursor:
            fn     = _vf(doc.get("display_name") or "")
            last   = _vf(doc.get("last_name",  ""))
            first  = _vf(doc.get("first_name", ""))
            middle = _vf(doc.get("middle_name",""))
            org    = _vf(doc.get("company_name",""))
            title  = _vf(doc.get("job_title",  ""))
            notes  = _vf(doc.get("notes",      ""))
            web    = _vf(doc.get("web_page",   ""))
            emails = doc.get("email_addresses") or [doc.get("from_addr") or ""]
            phones = doc.get("phone_numbers") or {}

            lines = ["BEGIN:VCARD", "VERSION:3.0"]
            if fn:
                lines.append(f"FN:{fn}")
                lines.append(f"N:{last};{first};{middle};;")
            for em in emails:
                if em:
                    lines.append(f"EMAIL;TYPE=INTERNET:{_vf(em)}")
            if org:   lines.append(f"ORG:{org}")
            if title: lines.append(f"TITLE:{title}")
            if phones.get("mobile"):   lines.append(f"TEL;TYPE=CELL:{_vf(phones['mobile'])}")
            if phones.get("business"): lines.append(f"TEL;TYPE=WORK:{_vf(phones['business'])}")
            if phones.get("home"):     lines.append(f"TEL;TYPE=HOME:{_vf(phones['home'])}")
            if web:   lines.append(f"URL:{web}")
            if notes: lines.append(f"NOTE:{notes}")

            # Business address
            bs = "; ".join(filter(None, [
                doc.get("business_street",""), doc.get("business_street2",""),
            ]))
            if any([bs, doc.get("business_city"), doc.get("business_state"),
                    doc.get("business_postal"), doc.get("business_country")]):
                lines.append(
                    f"ADR;TYPE=WORK:;;{_vf(bs)};{_vf(doc.get('business_city',''))};"
                    f"{_vf(doc.get('business_state',''))};{_vf(doc.get('business_postal',''))};"
                    f"{_vf(doc.get('business_country',''))}"
                )
            lines.append("END:VCARD")
            yield "\r\n".join(lines) + "\r\n\r\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/vcard; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="contacts_export.vcf"'},
    )


@app.route("/export/addresses.csv")
def export_addresses_csv():
    """Export the unique address book (from/to/cc across all emails) as CSV."""
    from email.utils import parseaddr as _parseaddr

    col     = get_col()
    q_txt   = request.args.get("q", "").strip().lower()
    sort_by = request.args.get("sort", "count")
    limit   = min(int(request.args.get("limit", 100_000)), 500_000)

    pipeline = [
        {"$project": {
            "date": 1,
            "addresses": {"$concatArrays": [
                [{"$ifNull": ["$from_addr", ""]}],
                {"$ifNull": ["$to_addrs", []]},
                {"$ifNull": ["$cc_addrs", []]},
            ]},
        }},
        {"$unwind": "$addresses"},
        {"$match": {"addresses": {"$nin": [None, ""]}}},
        {"$group": {
            "_id":        "$addresses",
            "count":      {"$sum": 1},
            "last_seen":  {"$max": "$date"},
            "first_seen": {"$min": "$date"},
        }},
        {"$sort": {"count": -1}},
        {"$limit": limit},
    ]

    raw = list(col.aggregate(pipeline, allowDiskUse=True))
    merged: dict = {}
    for row in raw:
        raw_addr = (row["_id"] or "").strip()
        if not raw_addr: continue
        name, addr = _parseaddr(raw_addr)
        if not addr or "@" not in addr:
            if "@" in raw_addr and " " not in raw_addr.strip("<> "):
                addr = raw_addr.strip("<> "); name = ""
            else: continue
        key = addr.lower()
        if key not in merged:
            merged[key] = {"email": key, "name": (name or "").strip(),
                           "count": 0, "last_seen": None, "first_seen": None}
        m = merged[key]
        m["count"] += row["count"]
        if not m["name"] and name: m["name"] = name.strip()
        ld, fd = row.get("last_seen"), row.get("first_seen")
        if ld and (m["last_seen"]  is None or ld > m["last_seen"]):  m["last_seen"]  = ld
        if fd and (m["first_seen"] is None or fd < m["first_seen"]): m["first_seen"] = fd

    entries = list(merged.values())
    if q_txt:
        entries = [e for e in entries if q_txt in e["email"] or q_txt in e["name"].lower()]
    if   sort_by == "name":      entries.sort(key=lambda x: x["name"].lower() or x["email"])
    elif sort_by == "email":     entries.sort(key=lambda x: x["email"])
    elif sort_by == "last_seen": entries.sort(key=lambda x: x["last_seen"] or datetime.datetime.min, reverse=True)
    else:                        entries.sort(key=lambda x: x["count"], reverse=True)

    headers = ["Email", "Display Name", "Times Seen", "First Seen", "Last Seen"]

    def rows():
        for e in entries:
            yield [
                e["email"], e["name"], e["count"],
                e["first_seen"].strftime("%Y-%m-%d") if e["first_seen"] else "",
                e["last_seen"].strftime("%Y-%m-%d")  if e["last_seen"]  else "",
            ]

    return Response(
        stream_with_context(_csv_stream(headers, rows())),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": 'attachment; filename="address_book_export.csv"'},
    )


@app.route("/export/attachments.csv")
def export_attachments_csv():
    """Export a manifest of every file in the Attachments folder as CSV."""
    folder  = request.args.get("folder", "all").strip()
    limit   = min(int(request.args.get("limit", 200_000)), 1_000_000)

    _ad   = _effective_attach_dir()
    roots = ([_ad] if folder == "all"
             else [os.path.join(_ad, folder)])

    headers = ["Folder", "Filename", "Size (bytes)", "Size", "Email Date"]

    # Build disk_path → date and filename → date maps from MongoDB.
    # Try the configured collection first, then fall back to the legacy local DB.
    _path_to_date:  dict = {}
    _fname_to_date: dict = {}

    def _fill_csv_date_maps(col):
        try:
            for row in col.aggregate([
                {"$match":   {"attachments": {"$exists": True}}},
                {"$unwind":  "$attachments"},
                {"$project": {"_id": 0,
                               "dp": "$attachments.disk_path",
                               "fn": "$attachments.filename",
                               "d":  "$date"}},
            ], allowDiskUse=True):
                d = row.get("d")
                if not d: continue
                dp = row.get("dp")
                fn = row.get("fn")
                if dp and dp not in _path_to_date:
                    _path_to_date[dp] = d
                elif dp and d < _path_to_date[dp]:
                    _path_to_date[dp] = d
                if fn and fn not in _fname_to_date:
                    _fname_to_date[fn] = d
                elif fn and d < _fname_to_date[fn]:
                    _fname_to_date[fn] = d
        except Exception:
            pass

    _fill_csv_date_maps(get_col())
    if not _path_to_date and not _fname_to_date:
        try:
            _fill_csv_date_maps(
                MongoClient("mongodb://localhost:27017",
                            serverSelectionTimeoutMS=1500)["mydb"]["pst_items"]
            )
        except Exception:
            pass

    def rows():
        count = 0
        for root in roots:
            if not os.path.isdir(root): continue
            for dirpath, _, filenames in os.walk(root):
                rel_dir = os.path.relpath(dirpath, _ad).replace(os.sep, "/")
                if rel_dir.split("/")[0] in _INTERNAL_FOLDERS: continue
                for fname in filenames:
                    if count >= limit: return
                    fpath = os.path.join(dirpath, fname)
                    try:
                        st  = os.stat(fpath)
                        sz  = st.st_size
                        if   sz < 1_024:         size_str = f"{sz} B"
                        elif sz < 1_048_576:     size_str = f"{sz/1_024:.1f} KB"
                        elif sz < 1_073_741_824: size_str = f"{sz/1_048_576:.1f} MB"
                        else:                    size_str = f"{sz/1_073_741_824:.2f} GB"
                        dt = _path_to_date.get(fpath) or _fname_to_date.get(fname)
                        if dt:
                            email_date = dt.strftime("%Y-%m-%d %H:%M")
                        else:
                            email_date = datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
                    except Exception:
                        sz, size_str, email_date = 0, "?", ""
                    yield [rel_dir, fname, sz, size_str, email_date]
                    count += 1

    return Response(
        stream_with_context(_csv_stream(headers, rows())),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": 'attachment; filename="attachments_manifest.csv"'},
    )


@app.route("/export/count")
def export_count():
    """Return record counts for the export UI (fast)."""
    col   = get_col()
    query, q = _build_record_query(request.args)
    if q:
        _ensure_text_index(col)
        query["$text"] = {"$search": q}
    return jsonify({"count": col.count_documents(query)})


# ---------------------------------------------------------------------------
# Timeline view
# ---------------------------------------------------------------------------

def _bucket_to_ts(bucket: str, granularity: str) -> int:
    """Return Unix timestamp (ms) for the midpoint of a bucket string."""
    try:
        if granularity == "year":
            dt = datetime.datetime(int(bucket), 7, 1)
        elif granularity == "quarter":
            year_s, q_s = bucket.split("-Q")
            month = (int(q_s) - 1) * 3 + 2   # mid-month of quarter
            dt = datetime.datetime(int(year_s), month, 15)
        elif granularity == "week":
            # "%G-W%V-4" = ISO Thursday of that week
            dt = datetime.datetime.strptime(bucket + "-4", "%G-W%V-%u")
        elif granularity == "day":
            dt = datetime.datetime.strptime(bucket, "%Y-%m-%d").replace(hour=12)
        else:  # month
            y, m = bucket.split("-")
            dt = datetime.datetime(int(y), int(m), 15)
        return int(dt.timestamp() * 1000)
    except Exception:
        return 0


@app.route("/timeline")
def timeline():
    """
    Return swim-lane timeline data: emails per (sender/folder, time-bucket).

    Query params:
      group_by    sender | folder  (default: sender)
      granularity day | week | month | quarter | year  (default: month)
      top_n       1-25  (default: 10)
      date_from / date_to  YYYY-MM-DD
      q           optional text filter
    """
    col         = get_col()
    group_by    = request.args.get("group_by",    "sender").strip().lower()
    granularity = request.args.get("granularity", "month" ).strip().lower()
    top_n       = min(max(1, int(request.args.get("top_n", 10))), 25)
    date_from   = request.args.get("date_from",  "").strip()
    date_to     = request.args.get("date_to",    "").strip()
    q           = request.args.get("q",          "").strip()

    if granularity not in ("day", "week", "month", "quarter", "year"):
        granularity = "month"

    # ── Base match ────────────────────────────────────────────────────────────
    base_match: dict = {"item_type": "email",
                        "date": {"$exists": True, "$type": "date"}}
    date_part: dict = {}
    if date_from:
        try:
            date_part["$gte"] = datetime.datetime.strptime(date_from, "%Y-%m-%d")
        except ValueError:
            pass
    if date_to:
        try:
            date_part["$lte"] = (datetime.datetime.strptime(date_to, "%Y-%m-%d")
                                 + datetime.timedelta(days=1))
        except ValueError:
            pass
    if date_part:
        base_match["date"].update(date_part)
    if q:
        _ensure_text_index(col)
        base_match["$text"] = {"$search": q}

    field = "folder_path" if group_by == "folder" else "from_addr"

    # ── Bucket expression ─────────────────────────────────────────────────────
    if granularity == "year":
        bucket_expr: dict = {"$dateToString": {"format": "%Y",      "date": "$date"}}
    elif granularity == "quarter":
        bucket_expr = {"$concat": [
            {"$toString": {"$year": "$date"}},
            "-Q",
            {"$toString": {"$toInt": {
                "$ceil": {"$divide": [{"$month": "$date"}, 3]}
            }}},
        ]}
    elif granularity == "week":
        bucket_expr = {"$dateToString": {"format": "%G-W%V", "date": "$date"}}
    elif granularity == "day":
        bucket_expr = {"$dateToString": {"format": "%Y-%m-%d", "date": "$date"}}
    else:  # month
        bucket_expr = {"$dateToString": {"format": "%Y-%m",   "date": "$date"}}

    # ── Top-N groups ──────────────────────────────────────────────────────────
    top_pipe = [
        {"$match": base_match},
        {"$group": {"_id": f"${field}", "n": {"$sum": 1}}},
        {"$sort": {"n": -1}},
        {"$limit": top_n},
    ]
    top_groups = [r["_id"] for r in col.aggregate(top_pipe, allowDiskUse=True)
                  if r["_id"]]

    if not top_groups:
        return jsonify({"groups": [], "buckets": [], "data": [],
                        "group_by": group_by, "granularity": granularity,
                        "field": field, "total_emails": 0})

    # ── Timeline aggregation ──────────────────────────────────────────────────
    tl_match = {**base_match, field: {"$in": top_groups}}
    tl_pipe  = [
        {"$match": tl_match},
        {"$group": {
            "_id":      {"group": f"${field}", "bucket": bucket_expr},
            "count":    {"$sum": 1},
        }},
        {"$sort": {"_id.bucket": 1}},
    ]
    raw = list(col.aggregate(tl_pipe, allowDiskUse=True))

    # ── Build data points and per-group stats ─────────────────────────────────
    data: list = []
    group_stats: dict = {}   # group → {total, first, last, peak_count, peak}

    for row in raw:
        group  = row["_id"]["group"]
        bucket = row["_id"]["bucket"]
        count  = row["count"]
        if not group or not bucket:
            continue
        bucket_ts = _bucket_to_ts(bucket, granularity)
        data.append({
            "group":     group,
            "group_idx": top_groups.index(group),
            "bucket":    bucket,
            "bucket_ts": bucket_ts,
            "count":     count,
        })
        gs = group_stats.setdefault(group, {
            "total": 0, "first": bucket, "last": bucket,
            "peak_count": 0, "peak": bucket,
        })
        gs["total"] += count
        if bucket < gs["first"]: gs["first"] = bucket
        if bucket > gs["last"]:  gs["last"]  = bucket
        if count > gs["peak_count"]:
            gs["peak_count"] = count
            gs["peak"]       = bucket

    # ordered summary rows (same order as top_groups)
    summary = []
    for g in top_groups:
        gs = group_stats.get(g, {})
        summary.append({
            "group":       g,
            "total":       gs.get("total", 0),
            "first":       gs.get("first", ""),
            "last":        gs.get("last",  ""),
            "peak":        gs.get("peak",  ""),
            "peak_count":  gs.get("peak_count", 0),
        })

    buckets = sorted({d["bucket"] for d in data})

    return jsonify({
        "groups":       top_groups,
        "buckets":      buckets,
        "data":         data,
        "summary":      summary,
        "group_by":     group_by,
        "granularity":  granularity,
        "field":        field,
        "total_emails": sum(d["count"] for d in data),
    })


# ---------------------------------------------------------------------------
# Analytics / Dashboard
# ---------------------------------------------------------------------------

@app.route("/analytics/volume")
def analytics_volume():
    """Email count grouped by year or year+month."""
    col         = get_col()
    granularity = request.args.get("granularity", "year").strip().lower()

    if granularity == "month":
        pipe = [
            {"$match": {"date": {"$nin": [None]}}},
            {"$group": {
                "_id":   {"year": {"$year": "$date"}, "month": {"$month": "$date"}},
                "count": {"$sum": 1},
            }},
            {"$sort": {"_id.year": 1, "_id.month": 1}},
        ]
        rows   = list(col.aggregate(pipe, allowDiskUse=True))
        labels = [f"{r['_id']['year']}-{r['_id']['month']:02d}" for r in rows if r["_id"]]
        counts = [r["count"] for r in rows if r["_id"]]
    else:
        pipe = [
            {"$match": {"date": {"$nin": [None]}}},
            {"$group": {
                "_id":   {"$year": "$date"},
                "count": {"$sum": 1},
            }},
            {"$sort": {"_id": 1}},
        ]
        rows   = list(col.aggregate(pipe, allowDiskUse=True))
        labels = [str(r["_id"]) for r in rows if r["_id"]]
        counts = [r["count"]    for r in rows if r["_id"]]

    return jsonify({"labels": labels, "counts": counts})


@app.route("/analytics/senders")
def analytics_senders():
    """Top N senders by email count."""
    col   = get_col()
    limit = min(int(request.args.get("limit", 15)), 100)

    pipe = [
        {"$match": {
            "item_type": "email",
            "from_addr": {"$nin": [None, ""]},
        }},
        {"$group": {
            "_id":   {"$toLower": "$from_addr"},
            "count": {"$sum": 1},
        }},
        {"$sort":  {"count": -1}},
        {"$limit": limit},
    ]
    rows = list(col.aggregate(pipe, allowDiskUse=True))
    return jsonify({
        "labels": [r["_id"]   for r in rows],
        "counts": [r["count"] for r in rows],
    })


@app.route("/analytics/attachment-types")
def analytics_attachment_types():
    """Count attachments by file extension (top N)."""
    col   = get_col()
    limit = min(int(request.args.get("limit", 12)), 30)

    pipe = [
        {"$match": {"has_attachments": True}},
        {"$unwind": "$attachments"},
        {"$project": {
            "ext": {
                "$toLower": {
                    "$let": {
                        "vars": {
                            "parts": {"$split": [
                                {"$ifNull": ["$attachments.filename", ""]}, "."
                            ]},
                        },
                        "in": {
                            "$cond": {
                                "if":   {"$gt": [{"$size": "$$parts"}, 1]},
                                "then": {"$arrayElemAt": ["$$parts", -1]},
                                "else": "",
                            }
                        },
                    }
                }
            }
        }},
        {"$match":  {"ext": {"$nin": [None, ""]}}},
        {"$group":  {"_id": "$ext", "count": {"$sum": 1}}},
        {"$sort":   {"count": -1}},
        {"$limit":  limit},
    ]
    rows = list(col.aggregate(pipe, allowDiskUse=True))
    return jsonify({
        "labels": [r["_id"]   for r in rows],
        "counts": [r["count"] for r in rows],
    })


@app.route("/analytics/folders")
def analytics_folders():
    """Top N folders by email count, with shortened display labels."""
    col   = get_col()
    limit = min(int(request.args.get("limit", 15)), 50)

    pipe = [
        {"$match": {"folder_path": {"$nin": [None, ""]}}},
        {"$group": {
            "_id":   "$folder_path",
            "count": {"$sum": 1},
        }},
        {"$sort":  {"count": -1}},
        {"$limit": limit},
    ]
    rows = list(col.aggregate(pipe, allowDiskUse=True))

    def _shorten(path, max_len=38):
        p = (path or "").replace("\\", "/")
        if len(p) <= max_len:
            return p
        parts = p.split("/")
        if len(parts) >= 2:
            return "…/" + "/".join(parts[-2:])
        return "…" + p[-max_len:]

    full_labels  = [r["_id"]          for r in rows]
    short_labels = [_shorten(r["_id"]) for r in rows]
    counts       = [r["count"]         for r in rows]

    return jsonify({"labels": short_labels, "full_labels": full_labels, "counts": counts})


# ---------------------------------------------------------------------------
# Bulk operations & tags
# ---------------------------------------------------------------------------

@app.route("/tags")
def list_tags():
    """Return all unique tag values in the collection (sorted)."""
    col  = get_col()
    tags = sorted(t for t in col.distinct("tags") if t)
    return jsonify(tags)


@app.route("/bulk/delete", methods=["POST"])
def bulk_delete():
    """Delete a list of records by _id.  Body: {"ids": [...]}"""
    data = request.get_json(silent=True) or {}
    ids  = data.get("ids", [])
    if not ids:
        return jsonify({"error": "No ids provided"}), 400
    col     = get_col()
    result  = col.delete_many({"_id": {"$in": ids}})
    return jsonify({"deleted": result.deleted_count})


@app.route("/bulk/tag", methods=["POST"])
def bulk_tag():
    """Add a tag to a list of records.  Body: {"ids": [...], "tag": "..."}"""
    data = request.get_json(silent=True) or {}
    ids  = data.get("ids", [])
    tag  = (data.get("tag") or "").strip()[:64]
    if not ids or not tag:
        return jsonify({"error": "ids and tag required"}), 400
    col    = get_col()
    result = col.update_many(
        {"_id": {"$in": ids}},
        {"$addToSet": {"tags": tag}},
    )
    return jsonify({"tagged": result.modified_count, "tag": tag})


@app.route("/bulk/untag", methods=["POST"])
def bulk_untag():
    """Remove a tag from a list of records.  Body: {"ids": [...], "tag": "..."}"""
    data = request.get_json(silent=True) or {}
    ids  = data.get("ids", [])
    tag  = (data.get("tag") or "").strip()
    if not ids or not tag:
        return jsonify({"error": "ids and tag required"}), 400
    col    = get_col()
    result = col.update_many(
        {"_id": {"$in": ids}},
        {"$pull": {"tags": tag}},
    )
    return jsonify({"untagged": result.modified_count, "tag": tag})


def _selected_docs(ids: list, col) -> list:
    """Fetch full export rows for the given IDs."""
    proj = {
        "date": 1, "from_addr": 1, "to_addrs": 1, "cc_addrs": 1,
        "subject": 1, "folder_path": 1, "item_type": 1,
        "has_attachments": 1, "body_plain": 1, "tags": 1,
    }
    return list(col.find({"_id": {"$in": ids}}, proj)
                   .sort("date", -1)
                   .limit(10_000))


@app.route("/export/selected.csv", methods=["POST"])
def export_selected_csv():
    """Stream a CSV of the POSTed list of record IDs."""
    data = request.get_json(silent=True) or {}
    ids  = data.get("ids", [])
    if not ids:
        return jsonify({"error": "No ids provided"}), 400
    col  = get_col()
    docs = _selected_docs(ids, col)

    COL_HEADERS = ["Date", "From", "To", "CC", "Subject",
                   "Folder", "Type", "Has Attachments", "Tags", "Body (first 500 chars)"]

    def rows():
        for doc in docs:
            body = " ".join((doc.get("body_plain") or "").split())[:500]
            yield [
                doc["date"].strftime("%Y-%m-%d %H:%M") if doc.get("date") else "",
                doc.get("from_addr")  or "",
                "; ".join(doc.get("to_addrs") or []),
                "; ".join(doc.get("cc_addrs") or []),
                doc.get("subject")    or "",
                doc.get("folder_path") or "",
                doc.get("item_type")  or "email",
                "Yes" if doc.get("has_attachments") else "No",
                "; ".join(doc.get("tags") or []),
                body,
            ]

    fname = f"selected_{len(ids)}_records.csv"
    return Response(
        stream_with_context(_csv_stream(COL_HEADERS, rows())),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.route("/export/selected.eml.zip", methods=["POST"])
def export_selected_eml_zip():
    """
    Package selected records as individual .eml files inside a ZIP archive.
    Body: {"ids": [...]}
    """
    import zipfile

    data = request.get_json(silent=True) or {}
    ids  = data.get("ids", [])
    if not ids:
        return jsonify({"error": "No ids provided"}), 400

    col  = get_col()
    proj = {
        "date": 1, "from_addr": 1, "to_addrs": 1, "cc_addrs": 1,
        "subject": 1, "folder_path": 1, "body_plain": 1, "body_html": 1,
        "message_id": 1,
    }
    docs = list(col.find({"_id": {"$in": ids}}, proj).sort("date", -1).limit(10_000))

    buf = io.BytesIO()
    seen_names: dict = {}   # deduplicate filenames inside the archive

    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for i, doc in enumerate(docs, 1):
            base = _eml_filename(doc)
            # Ensure uniqueness within the archive
            name = base
            if name in seen_names:
                stem, ext = name.rsplit(".", 1)
                name = f"{stem}_{i}.{ext}"
            seen_names[name] = True

            try:
                eml_bytes = _build_eml(doc)
            except Exception:
                eml_bytes = b""   # skip broken records silently

            zf.writestr(name, eml_bytes)

    buf.seek(0)
    fname = f"selected_{len(docs)}_emails.zip"
    return Response(
        buf.read(),
        mimetype="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.route("/export/selected.xlsx", methods=["POST"])
def export_selected_xlsx():
    """Export the POSTed list of record IDs as a formatted Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    data = request.get_json(silent=True) or {}
    ids  = data.get("ids", [])
    if not ids:
        return jsonify({"error": "No ids provided"}), 400
    col  = get_col()
    docs = _selected_docs(ids, col)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Selected Records"

    HEADERS = ["Date", "From", "To", "CC", "Subject",
               "Folder", "Type", "Has Attachments", "Tags", "Body (first 500 chars)"]
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1A1F36")
    hdr_align = Alignment(horizontal="left", vertical="center")
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
    ws.freeze_panes = "A2"

    for doc in docs:
        body = " ".join((doc.get("body_plain") or "").split())[:500]
        ws.append([
            doc["date"].strftime("%Y-%m-%d %H:%M") if doc.get("date") else "",
            doc.get("from_addr")   or "",
            "; ".join(doc.get("to_addrs") or []),
            "; ".join(doc.get("cc_addrs") or []),
            doc.get("subject")     or "",
            doc.get("folder_path") or "",
            doc.get("item_type")   or "email",
            "Yes" if doc.get("has_attachments") else "No",
            "; ".join(doc.get("tags") or []),
            body,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"selected_{len(ids)}_records.xlsx"
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# Folder Tree
# ---------------------------------------------------------------------------

def _build_folder_tree(path_counts: dict) -> list:
    """
    Build a nested folder tree from {folder_path: count}.
    Paths are normalised to use '/' as separator.
    Returns a sorted list of root-level node dicts:
        {name, path, own_count, total_count, children: [...]}
    """
    # Normalise paths
    normalised: dict = {}
    for raw, cnt in path_counts.items():
        path = (raw or "").replace("\\", "/").strip("/")
        if path:
            normalised[path] = normalised.get(path, 0) + cnt

    # Build tree in-place as nested dicts  {segment -> node_dict}
    roots: dict = {}

    def _get_or_create(parent_map: dict, seg: str) -> dict:
        if seg not in parent_map:
            parent_map[seg] = {"name": seg, "path": "", "own_count": 0, "children": {}}
        return parent_map[seg]

    for path, cnt in normalised.items():
        segments = [s for s in path.split("/") if s]
        node_map = roots
        for seg in segments:
            node = _get_or_create(node_map, seg)
            node_map = node["children"]
        node["own_count"] += cnt  # last segment owns the count

    # Assign full paths and compute total_count, convert children dicts→lists
    def _finalise(nodes: dict, prefix: str) -> list:
        result = []
        for name in sorted(nodes.keys(), key=str.lower):
            node      = nodes[name]
            node["path"] = (prefix + "/" + name) if prefix else name
            children  = _finalise(node["children"], node["path"])
            total     = node["own_count"] + sum(c["total_count"] for c in children)
            result.append({
                "name":        name,
                "path":        node["path"],
                "own_count":   node["own_count"],
                "total_count": total,
                "children":    children,
            })
        return result

    return _finalise(roots, "")


@app.route("/folder-tree")
def folder_tree():
    """Return the PST folder hierarchy with per-folder email counts."""
    col      = get_col()
    pipeline = [
        {"$match": {"folder_path": {"$exists": True, "$nin": [None, ""]}}},
        {"$group": {"_id": "$folder_path", "count": {"$sum": 1}}},
    ]
    raw           = {doc["_id"]: doc["count"] for doc in col.aggregate(pipeline) if doc.get("_id")}
    tree          = _build_folder_tree(raw)
    total_folders = len(raw)
    total_emails  = sum(raw.values())
    return jsonify({"tree": tree, "total_folders": total_folders, "total_emails": total_emails})


# ---------------------------------------------------------------------------
# Duplicate Detection
# ---------------------------------------------------------------------------

def _folder_priority(folder_path: str) -> int:
    """Lower number = more important folder → preferred to keep."""
    fp = (folder_path or "").lower()
    if "inbox"   in fp: return 0
    if "sent"    in fp: return 1
    if "draft"   in fp: return 3
    if "delete"  in fp or "trash" in fp: return 8
    if "junk"    in fp or "spam"  in fp: return 9
    return 5


@app.route("/duplicates")
def find_duplicates():
    """
    Find email records that appear more than once.
    Strategy A: group by message_id (exact duplicate).
    Strategy B: group by (subject_lower + from_lower + date_day) for emails
                that have no message_id.
    Returns paginated list of groups, each with which _id to keep.
    """
    col      = get_col()
    page     = max(1, int(request.args.get("page",     1)))
    per_page = min(int(request.args.get("per_page",   50)), 200)
    strategy = request.args.get("strategy", "both").strip().lower()  # both | msgid | content

    results = []

    # ── Strategy A: message_id ────────────────────────────────────────────────
    if strategy in ("both", "msgid"):
     pipe_a = [
        {"$match": {
            "item_type":  "email",
            "message_id": {"$exists": True, "$nin": [None, ""]},
        }},
        {"$group": {
            "_id":          "$message_id",
            "count":        {"$sum": 1},
            "record_ids":   {"$push": "$_id"},
            "subject":      {"$first": "$subject"},
            "from_addr":    {"$first": "$from_addr"},
            "earliest":     {"$min": "$date"},
            "folder_paths": {"$addToSet": "$folder_path"},
        }},
        {"$match": {"count": {"$gt": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 20_000},
     ]
     for row in col.aggregate(pipe_a, allowDiskUse=True):
        results.append({
            "key":          row["_id"],
            "method":       "message_id",
            "count":        row["count"],
            "record_ids":   [str(i) for i in row["record_ids"]],
            "subject":      row.get("subject") or "(no subject)",
            "from_addr":    row.get("from_addr") or "",
            "earliest":     row["earliest"].strftime("%Y-%m-%d") if row.get("earliest") else "",
            "folder_paths": [f for f in (row.get("folder_paths") or []) if f],
            "duplicates":   row["count"] - 1,
        })

    # ── Strategy B: content fingerprint (no message_id) ───────────────────────
    if strategy in ("both", "content"):
     pipe_b = [
        {"$match": {
            "item_type": "email",
            "$or": [{"message_id": {"$exists": False}},
                    {"message_id": {"$in": [None, ""]}}],
        }},
        {"$addFields": {
            "_day": {"$cond": {
                "if":   {"$gt": ["$date", None]},
                "then": {"$dateToString": {"format": "%Y-%m-%d", "date": "$date"}},
                "else": "nodate",
            }},
        }},
        {"$group": {
            "_id": {
                "subj": {"$toLower": {"$ifNull": ["$subject", ""]}},
                "from": {"$toLower": {"$ifNull": ["$from_addr", ""]}},
                "day":  "$_day",
            },
            "count":        {"$sum": 1},
            "record_ids":   {"$push": "$_id"},
            "subject":      {"$first": "$subject"},
            "from_addr":    {"$first": "$from_addr"},
            "earliest":     {"$min": "$date"},
            "folder_paths": {"$addToSet": "$folder_path"},
        }},
        {"$match": {"count": {"$gt": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10_000},
     ]
     for row in col.aggregate(pipe_b, allowDiskUse=True):
        gid = row["_id"]
        results.append({
            "key":          f"{gid['subj']}|{gid['from']}|{gid['day']}",
            "method":       "content",
            "count":        row["count"],
            "record_ids":   [str(i) for i in row["record_ids"]],
            "subject":      row.get("subject") or "(no subject)",
            "from_addr":    row.get("from_addr") or "",
            "earliest":     row["earliest"].strftime("%Y-%m-%d") if row.get("earliest") else "",
            "folder_paths": [f for f in (row.get("folder_paths") or []) if f],
            "duplicates":   row["count"] - 1,
        })

    total       = len(results)
    extra_total = sum(r["duplicates"] for r in results)
    msgid_count = sum(1 for r in results if r["method"] == "message_id")
    cont_count  = total - msgid_count

    start     = (page - 1) * per_page
    page_data = results[start: start + per_page]

    return jsonify({
        "groups":       page_data,
        "total":        total,
        "extra_total":  extra_total,
        "msgid_groups": msgid_count,
        "cont_groups":  cont_count,
        "page":         page,
        "per_page":     per_page,
        "pages":        max(1, -(-total // per_page)),
    })


@app.route("/duplicates/delete", methods=["POST"])
def delete_duplicates():
    """
    Given a list of record _ids (all copies in a group), keep the best one
    (highest-priority folder then oldest date) and delete the rest.
    Accepts: {ids: [...]} or {delete_ids: [...]} to delete specific IDs directly.
    Returns: {deleted: N, kept: id}
    """
    data = request.get_json(silent=True) or {}
    col  = get_col()

    # Caller can supply either all IDs (server picks which to keep) …
    all_ids    = data.get("ids",        [])
    # … or just the IDs to delete (caller already decided which to keep)
    delete_ids = data.get("delete_ids", [])

    if delete_ids:
        result = col.delete_many({"_id": {"$in": delete_ids}})
        return jsonify({"deleted": result.deleted_count, "kept": None})

    if not all_ids or len(all_ids) < 2:
        return jsonify({"error": "Need at least 2 IDs"}), 400

    docs = list(col.find(
        {"_id": {"$in": all_ids}},
        {"_id": 1, "date": 1, "folder_path": 1},
    ))
    if len(docs) < 2:
        return jsonify({"deleted": 0, "kept": str(docs[0]["_id"]) if docs else None})

    docs.sort(key=lambda d: (
        _folder_priority(d.get("folder_path")),
        d.get("date") or datetime.datetime.max,
    ))
    keep_id   = docs[0]["_id"]
    to_delete = [d["_id"] for d in docs[1:]]

    result = col.delete_many({"_id": {"$in": to_delete}})
    return jsonify({"deleted": result.deleted_count, "kept": str(keep_id)})


# ---------------------------------------------------------------------------
# OCR Dashboard
# ---------------------------------------------------------------------------

SEARCHABLE_CHARS_THRESHOLD = 100   # chars below this → "scanned" (image-based) PDF


@app.route("/ocr-pending")
def ocr_pending():
    """Return count of PDFs not yet OCR'd — used by navbar badge."""
    _pdf_dir = get_pdf_dir(); _pdf_txt = get_pdf_text_dir()
    if not os.path.isdir(_pdf_dir):
        return jsonify({"pending": 0, "total": 0})
    pdf_files = [f for f in os.listdir(_pdf_dir) if f.lower().endswith(".pdf")]
    total = len(pdf_files)
    pending = sum(
        1 for f in pdf_files
        if not os.path.isfile(os.path.join(_pdf_txt, os.path.splitext(f)[0] + ".txt"))
    )
    return jsonify({"pending": pending, "total": total})


@app.route("/ocr-status")
def ocr_status():
    """
    List every PDF in Attachments/pdf/ with its OCR status.
    Returns: { pdfs: [...], summary: {total, searchable, scanned, not_indexed} }
    """
    _pdf_dir = get_pdf_dir(); _pdf_txt = get_pdf_text_dir()
    if not os.path.isdir(_pdf_dir):
        return jsonify({"pdfs": [], "summary": {
            "total": 0, "searchable": 0, "scanned": 0, "not_indexed": 0}})

    pdf_files = sorted(
        f for f in os.listdir(_pdf_dir) if f.lower().endswith(".pdf")
    )

    results = []
    summary = {"total": len(pdf_files), "searchable": 0, "scanned": 0, "not_indexed": 0}

    for fname in pdf_files:
        fpath      = os.path.join(_pdf_dir, fname)
        cache_name = os.path.splitext(fname)[0] + ".txt"
        cache_path = os.path.join(_pdf_txt, cache_name)

        try:
            sz = os.path.getsize(fpath)
            if   sz < 1_024:         size_str = f"{sz} B"
            elif sz < 1_048_576:     size_str = f"{sz/1_024:.1f} KB"
            elif sz < 1_073_741_824: size_str = f"{sz/1_048_576:.1f} MB"
            else:                    size_str = f"{sz/1_073_741_824:.2f} GB"
        except Exception:
            sz, size_str = 0, "?"

        if os.path.isfile(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8", errors="replace") as fh:
                    text = fh.read(100_000)
                char_count = len(text.strip())
            except Exception:
                char_count = 0

            if char_count >= SEARCHABLE_CHARS_THRESHOLD:
                status = "searchable";  summary["searchable"] += 1
            else:
                status = "scanned";     summary["scanned"]    += 1
        else:
            char_count = None
            status = "not_indexed";     summary["not_indexed"] += 1

        results.append({
            "filename":   fname,
            "size_bytes": sz,
            "size_str":   size_str,
            "status":     status,
            "char_count": char_count,
        })

    return jsonify({"pdfs": results, "summary": summary})


def _run_ocr_job(job_id: str, filenames: list, pdf_dir: str, pdf_text_dir: str, q: queue.Queue, user_email: str = ""):
    """
    Background thread: run OCR on the listed PDFs, update the text cache.
    Emits JSON progress events: {msg, pct, ok, done}.
    """
    import shutil

    def emit(msg, pct, ok=True, done=False):
        q.put(json.dumps({"msg": msg, "pct": pct, "ok": ok, "done": done}))

    total = len(filenames)
    if total == 0:
        emit("No files to process.", 100, done=True)
        q.put(None)
        return

    # Check for ocrmypdf CLI — also look in the active venv's Scripts folder
    ocrmypdf_cmd = shutil.which("ocrmypdf")
    if not ocrmypdf_cmd:
        # Flask may be running inside a venv whose Scripts dir isn't on PATH
        venv_scripts = os.path.dirname(sys.executable)
        for candidate in ("ocrmypdf.exe", "ocrmypdf"):
            p = os.path.join(venv_scripts, candidate)
            if os.path.isfile(p):
                ocrmypdf_cmd = p
                break
    if not ocrmypdf_cmd:
        emit(
            "ocrmypdf not found on PATH. "
            "Install it with:  pip install ocrmypdf  (also needs Tesseract).",
            100, ok=False, done=True,
        )
        jobs[job_id]["status"] = "error"
        q.put(None)
        return

    os.makedirs(pdf_text_dir, exist_ok=True)

    for i, fname in enumerate(filenames, 1):
        base_pct   = int((i - 1) / total * 94)
        cache_name = os.path.splitext(fname)[0] + ".txt"
        cache_path = os.path.join(pdf_text_dir, cache_name)
        src_path   = os.path.join(pdf_dir, fname)
        tmp_path   = src_path + ".__ocr_tmp__.pdf"

        emit(f"[{i}/{total}] {fname[:60]}", base_pct)

        if not os.path.isfile(src_path):
            emit(f"  Skipped — file not found: {fname}", base_pct)
            continue

        try:
            proc = subprocess.run(
                [ocrmypdf_cmd, "--skip-text", "--quiet", src_path, tmp_path],
                capture_output=True, text=True, timeout=300,
            )
            # ocrmypdf exit code 6 means "already has text, skipped" which is fine
            if proc.returncode not in (0, 6):
                stderr_short = (proc.stderr or "").strip()[:120]
                emit(f"  OCR warning (rc={proc.returncode}): {stderr_short}", base_pct)

            # Extract text from the newly OCR'd PDF (or original if OCR was skipped)
            source = tmp_path if os.path.isfile(tmp_path) else src_path
            text   = _extract_file_text(source)

            with open(cache_path, "w", encoding="utf-8", errors="replace") as fh:
                fh.write(text)

            char_count = len(text.strip())
            emit(f"  ✓ {char_count:,} chars extracted", base_pct + 1)

        except subprocess.TimeoutExpired:
            emit(f"  Timed-out after 5 min: {fname}", base_pct, ok=False)
        except Exception as e:
            emit(f"  Error: {e}", base_pct, ok=False)
        finally:
            if os.path.isfile(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

    emit(f"OCR complete — {total} file(s) processed.", 100, done=True)
    jobs[job_id]["status"] = "done"
    _send_notification_email(
        user_email,
        "✅ Attachment OCR complete — pstbrowser.com",
        f"OCR processing is complete for {total} PDF(s). You can now search within your documents at https://pstbrowser.com",
        f"<p>OCR processing is complete for <strong>{total} PDF(s)</strong>.</p>"
        f"<p>You can now search inside your PDF attachments — not just filenames, but the full content of each document.</p>"
        f"<p><a href='https://pstbrowser.com/?tab=browse' style='background:#2563eb;color:#fff;"
        f"padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:600;"
        f"display:inline-block'>Browse and search attachments</a></p>",
    )
    q.put(None)


@app.route("/ocr-run", methods=["POST"])
def ocr_run():
    """Start an OCR background job. Body: {files: ['name.pdf', ...]}. Returns {job_id}."""
    data      = request.get_json(silent=True) or {}
    filenames = data.get("files", [])

    # Security: only allow basenames of .pdf files — no path components
    filenames = [
        os.path.basename(f) for f in filenames
        if f.lower().endswith(".pdf")
    ]
    if not filenames:
        return jsonify({"error": "No valid PDF filenames provided"}), 400

    # Capture dirs before thread (no request context inside thread)
    _pdf_dir     = get_pdf_dir()
    _pdf_txt_dir = get_pdf_text_dir()

    job_id     = str(uuid.uuid4())
    q          = queue.Queue()
    user_email = current_user.email if current_user.is_authenticated else ""
    jobs[job_id] = {
        "queue":    q,
        "status":   "running",
        "filename": f"{len(filenames)} PDF(s)",
    }

    threading.Thread(
        target=_run_ocr_job, args=(job_id, filenames, _pdf_dir, _pdf_txt_dir, q, user_email),
        daemon=True,
    ).start()
    return jsonify({"job_id": job_id, "count": len(filenames)})


@app.route("/ocr-progress/<job_id>")
def ocr_progress(job_id: str):
    """SSE stream emitting JSON progress events for an OCR job."""
    if job_id not in jobs:
        return jsonify({"error": "Job not found"}), 404

    def generate():
        q = jobs[job_id]["queue"]
        while True:
            try:
                item = q.get(timeout=60)
            except queue.Empty:
                yield "data: \n\n"   # keep-alive
                continue
            if item is None:
                break
            yield f"data: {item}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True, port=5000, threaded=True)
